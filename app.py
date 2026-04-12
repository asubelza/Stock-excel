import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

EXCEL_FILE = "planilla_base.xlsx"

LIGHT_THEME = {
    'bg': '#F5F7FA', 'fg': '#2C3E50', 'accent': '#3498DB',
    'toolbar': '#FFFFFF', 'tree_bg': '#FFFFFF', 'tree_fg': '#2C3E50',
    'tree_sel': '#3498DB', 'entry_bg': '#FFFFFF', 'entry_fg': '#2C3E50',
    'button': '#3498DB', 'watermark': '#7F8C8D', 'warning': '#E74C3C', 'success': '#27AE60'
}

DARK_THEME = {
    'bg': '#1A1A2E', 'fg': '#EAEAEA', 'accent': '#4A90D9',
    'toolbar': '#16213E', 'tree_bg': '#16213E', 'tree_fg': '#EAEAEA',
    'tree_sel': '#4A90D9', 'entry_bg': '#252545', 'entry_fg': '#EAEAEA',
    'button': '#4A90D9', 'watermark': '#6C7A89', 'warning': '#E74C3C', 'success': '#2ECC71'
}

USUARIOS_DEFAULT = [
    {'user': 'admin', 'pass': 'admin123', 'nombre': 'Administrador', 'rol': 'admin'},
    {'user': 'deposito', 'pass': 'depo123', 'nombre': 'Deposito', 'rol': 'user'},
    {'user': 'ventas', 'pass': 'vta123', 'nombre': 'Ventas', 'rol': 'user'},
]

class StockApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestion de Stock")
        self.root.geometry("1100x750")
        
        self.wb = None
        self.ws_productos = None
        self.ws_movimientos = None
        self.ws_usuarios = None
        self.ws_proveedores = None
        self.products = []
        self.depositos = ['Principal']
        self.dark_mode = False
        self.current_theme = LIGHT_THEME
        self.usuario_actual = None
        self.status_label = None
        
        self.init_excel()
        self.login()
        
        if self.usuario_actual:
            self.setup_styles()
            self.setup_ui()
            self.actualizar_status()
            self.filter_products()
    
    def login(self):
        win = tk.Toplevel(self.root)
        win.title("Login")
        win.geometry("300x220")
        win.transient(self.root)
        win.grab_set()
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        ttk.Label(win, text="Gestion de Stock", font=('Arial', 14, 'bold')).pack(pady=20)
        ttk.Label(win, text="Usuario:").pack(pady=5)
        user_var = tk.StringVar()
        ttk.Entry(win, textvariable=user_var, width=20).pack()
        ttk.Label(win, text="Contrasena:").pack(pady=5)
        pass_var = tk.StringVar()
        ttk.Entry(win, textvariable=pass_var, show="*", width=20).pack()
        
        def entrar():
            user = user_var.get().strip()
            password = pass_var.get().strip()
            usuarios = self.get_usuarios()
            for u in usuarios:
                if u['user'] == user and u['pass'] == password:
                    self.usuario_actual = u
                    win.destroy()
                    return
            messagebox.showerror("Error", "Usuario o contrasena incorrectos")
        
        ttk.Button(win, text="Ingresar", command=entrar).pack(pady=20)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8), fg=theme['watermark'], bg=theme['bg']).pack()
        win.protocol("WM_DELETE_WINDOW", lambda: self.root.destroy())
        self.root.wait_window(win)
    
    def get_proveedores(self):
        try:
            sheet_names = self.wb.sheetnames
            if 'Proveedores' not in sheet_names:
                self.ws_proveedores = self.wb.create_sheet('Proveedores')
                headers = ['cuit', 'nombre', 'direccion', 'telefono', 'email']
                for col, h in enumerate(headers, 1):
                    self.ws_proveedores.cell(1, col).value = h
                self.wb.save(EXCEL_FILE)
            else:
                self.ws_proveedores = self.wb['Proveedores']
            
            proveedores = []
            for row in range(2, self.ws_proveedores.max_row + 1):
                cuit = self.ws_proveedores.cell(row, 1).value
                if cuit:
                    proveedores.append({
                        'cuit': str(cuit),
                        'nombre': self.ws_proveedores.cell(row, 2).value,
                        'direccion': self.ws_proveedores.cell(row, 3).value,
                        'telefono': self.ws_proveedores.cell(row, 4).value,
                        'email': self.ws_proveedores.cell(row, 5).value,
                    })
            return proveedores
        except:
            return []
    
    def buscar_arca(self, cuits):
        import urllib.request
        import json
        import ssl
        
        cuits = cuits.strip().replace("-", "").replace(" ", "")
        
        if len(cuits) < 7:
            return None
        
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            url = f"https://www.afip.gob.ar/rcel/consultaPadron/padron.aspx?cuitEmpresa=&razonSocial=&domicilio=&provincia=&codigo Actividad=&ano=&mes="
            
            data = {
                "cuit": cuits,
                "token": ""
            }
            
            req = urllib.request.Request(
                "https://soa.afip.gob.ar/wscthomo/v1/contribuyentes",
                data=json.dumps(data).encode('utf-8'),
                headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                method='POST'
            )
            
            with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
                result = json.loads(response.read().decode())
                if result.get('data'):
                    d = result['data']
                    return {
                        'cuit': cuits,
                        'nombre': d.get('razonSocial', 'Sin nombre'),
                        'direccion': d.get('domicilio', ''),
                        'telefono': d.get('telefono', ''),
                        'email': d.get('email', '')
                    }
        except Exception as e:
            print(f"ARCA error: {e}")
        
        try:
            import requests
            url = f"https://api-dolar-v1.vercel.app/info/{cuits}"
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                result = resp.json()
                if result.get('cuit'):
                    return {
                        'cuit': cuits,
                        'nombre': result.get('denominacion', 'Sin nombre'),
                        'direccion': result.get('domicilio', ''),
                        'telefono': '',
                        'email': ''
                    }
        except:
            pass
        
        return None
    
    def guardar_proveedor(self, proveedor_data):
        try:
            sheet_names = self.wb.sheetnames
            if 'Proveedores' not in sheet_names:
                self.ws_proveedores = self.wb.create_sheet('Proveedores')
                headers = ['cuit', 'nombre', 'direccion', 'telefono', 'email']
                for col, h in enumerate(headers, 1):
                    self.ws_proveedores.cell(1, col).value = h
            else:
                self.ws_proveedores = self.wb['Proveedores']
            
            row = self.ws_proveedores.max_row + 1
            self.ws_proveedores.cell(row, 1).value = proveedor_data.get('cuit', '')
            self.ws_proveedores.cell(row, 2).value = proveedor_data.get('nombre', '')
            self.ws_proveedores.cell(row, 3).value = proveedor_data.get('direccion', '')
            self.ws_proveedores.cell(row, 4).value = proveedor_data.get('telefono', '')
            self.ws_proveedores.cell(row, 5).value = proveedor_data.get('email', '')
            
            self.wb.save(EXCEL_FILE)
            return True
        except:
            return False
    
    def get_usuarios(self):
        try:
            sheet_names = self.wb.sheetnames
            if 'Usuarios' not in sheet_names:
                self.ws_usuarios = self.wb.create_sheet('Usuarios')
                headers = ['user', 'pass', 'nombre', 'rol']
                for col, h in enumerate(headers, 1):
                    self.ws_usuarios.cell(1, col).value = h
                for row, u in enumerate(USUARIOS_DEFAULT, 2):
                    self.ws_usuarios.cell(row, 1).value = u['user']
                    self.ws_usuarios.cell(row, 2).value = u['pass']
                    self.ws_usuarios.cell(row, 3).value = u['nombre']
                    self.ws_usuarios.cell(row, 4).value = u['rol']
                self.wb.save(EXCEL_FILE)
            else:
                self.ws_usuarios = self.wb['Usuarios']
            
            usuarios = []
            for row in range(2, self.ws_usuarios.max_row + 1):
                user = self.ws_usuarios.cell(row, 1).value
                if user:
                    usuarios.append({
                        'user': user,
                        'pass': self.ws_usuarios.cell(row, 2).value,
                        'nombre': self.ws_usuarios.cell(row, 3).value,
                        'rol': self.ws_usuarios.cell(row, 4).value,
                    })
            return usuarios
        except:
            return USUARIOS_DEFAULT
    
    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.update_theme()
    
    def update_theme(self):
        theme = DARK_THEME if self.dark_mode else LIGHT_THEME
        self.current_theme = theme
        self.root.configure(bg=theme['bg'])
        
        self.style.configure('TFrame', background=theme['bg'])
        self.style.configure('TLabel', background=theme['bg'], foreground=theme['fg'])
        self.style.configure('Treeview', background=theme['tree_bg'], foreground=theme['tree_fg'], fieldbackground=theme['tree_bg'])
        self.style.configure('Treeview.Heading', background=theme['toolbar'], foreground=theme['fg'])
        self.style.map('Treeview', background=[('selected', theme['tree_sel'])])
        self.style.configure('TButton', background=theme['button'], foreground='#FFFFFF')
        
        if hasattr(self, 'watermark'):
            self.watermark.configure(fg=theme['watermark'], bg=theme['bg'])
    
    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.update_theme()
    
    def init_excel(self):
        if not os.path.exists(EXCEL_FILE):
            messagebox.showerror("Error", f"No encontrado: {EXCEL_FILE}")
            return
        
        self.wb = load_workbook(EXCEL_FILE)
        sheet_names = self.wb.sheetnames
        
        if 'Movimientos' not in sheet_names:
            self.ws_movimientos = self.wb.create_sheet('Movimientos')
            headers = ['Fecha', 'Usuario', 'SKU', 'Producto', 'Tipo', 'Cantidad', 'Deposito', 'Nro-Comprobante', 'Nro-Factura', 'Nota', 'Costo', 'Lote']
            for col, h in enumerate(headers, 1):
                self.ws_movimientos.cell(1, col).value = h
        else:
            self.ws_movimientos = self.wb['Movimientos']
        
        self.ws_productos = self.wb.active
        
        self.products = []
        for row in range(4, self.ws_productos.max_row + 1):
            nombre = self.ws_productos.cell(row, 2).value
            sku = self.ws_productos.cell(row, 3).value
            if sku or nombre:
                stock = self.ws_productos.cell(row, 25).value or 0
                deposito = self.ws_productos.cell(row, 24).value or 'Principal'
                if deposito not in self.depositos:
                    self.depositos.append(deposito)
                self.products.append({
                    'row': row, 'Nombre': nombre, 'SKU': sku,
                    'Tipo': self.ws_productos.cell(row, 4).value,
                    'Estado': self.ws_productos.cell(row, 5).value,
                    'Rubro': self.ws_productos.cell(row, 8).value,
                    'Stock Min': self.ws_productos.cell(row, 21).value or 0,
                    'stock': stock, 'deposito': deposito,
                    'precio': self.ws_productos.cell(row, 16).value,
                })
    
    def actualizar_status(self):
        if self.status_label and self.usuario_actual:
            self.status_label.config(text=f"Excel: {EXCEL_FILE} | Usuario: {self.usuario_actual['nombre']}")
    
    def setup_ui(self):
        main = ttk.Frame(self.root, padding="10")
        main.pack(fill=tk.BOTH, expand=True)
        
        header = ttk.Frame(main)
        header.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(header, text="Gestion de Inventario", font=('Arial', 16, 'bold')).pack(side=tk.LEFT)
        
        toolbar = ttk.Frame(main)
        toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(toolbar, text="+ Producto", command=self.add_product).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Editar", command=self.edit_product).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Eliminar", command=self.delete_product).pack(side=tk.LEFT, padx=2)
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Button(toolbar, text="Entrada", command=self.entrada_stock).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Salida", command=self.salida_stock).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Transferir", command=self.transferir).pack(side=tk.LEFT, padx=2)
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Button(toolbar, text="Reportes", command=self.reportes).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Stock Bajo", command=self.stock_bajo).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Historial", command=self.historial).pack(side=tk.LEFT, padx=2)
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Button(toolbar, text="Guardar", command=self.guardar).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Modo Oscuro" if not self.dark_mode else "Modo Claro", command=self.toggle_theme).pack(side=tk.LEFT, padx=2)
        
        filters = ttk.Frame(main)
        filters.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(filters, text="Buscar:").pack(side=tk.LEFT)
        self.buscar_var = tk.StringVar()
        self.buscar_var.trace('w', lambda *a: self.on_search_change())
        buscar_entry = ttk.Entry(filters, textvariable=self.buscar_var, width=20)
        buscar_entry.pack(side=tk.LEFT, padx=5)
        buscar_entry.bind('<Return>', self.selectFirstMatch)
        ttk.Button(filters, text="Buscar", command=self.selectFirstMatch, width=6).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(filters, text="Deposito:").pack(side=tk.LEFT, padx=(15, 0))
        self.deposito_var = tk.StringVar(value='Principal')
        self.deposito_combo = ttk.Combobox(filters, textvariable=self.deposito_var, values=self.depositos, state='readonly', width=12)
        self.deposito_combo.pack(side=tk.LEFT, padx=5)
        self.deposito_combo.bind('<<ComboboxSelected>>', lambda e: self.filter_products())
        
        ttk.Label(filters, text="Rubro:").pack(side=tk.LEFT, padx=(15, 0))
        self.rubro_var = tk.StringVar()
        self.rubro_combo = ttk.Combobox(filters, textvariable=self.rubro_var, state='readonly', width=12)
        self.rubro_combo.pack(side=tk.LEFT, padx=5)
        self.rubro_combo.bind('<<ComboboxSelected>>', lambda e: self.filter_products())
        
        tree_frame = ttk.Frame(main)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        cols = ('SKU', 'Nombre', 'Stock', 'Stock Min', 'Deposito', 'Precio', 'Estado')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100 if col != 'Nombre' else 220)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        
        self.tree.tag_configure('bajo', foreground=self.current_theme['warning'])
        self.tree.tag_configure('ok', foreground=self.current_theme['success'])
        
        status_frame = ttk.Frame(main)
        status_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.status_label = ttk.Label(status_frame, text="")
        self.status_label.pack(side=tk.LEFT)
        
        ttk.Label(status_frame, text="|").pack(side=tk.LEFT, padx=10)
        ttk.Label(status_frame, text="Total:").pack(side=tk.LEFT)
        self.total_label = ttk.Label(status_frame, text="0")
        self.total_label.pack(side=tk.LEFT)
        
        ttk.Label(status_frame, text="|").pack(side=tk.LEFT, padx=10)
        ttk.Label(status_frame, text="Stock Bajo:").pack(side=tk.LEFT)
        self.bajo_label = ttk.Label(status_frame, text="0", foreground=self.current_theme['warning'])
        self.bajo_label.pack(side=tk.LEFT)
        
        self.watermark = tk.Label(self.root, text="Desarrollado por asubelzacg", font=('Arial', 9, 'bold'),
                            fg=LIGHT_THEME['watermark'], bg=LIGHT_THEME['bg'])
        self.watermark.place(relx=0.5, rely=0.99, anchor='center')
    
    def on_search_change(self):
        query = self.buscar_var.get().strip().lower()
        
        if len(query) >= 2:
            matches = []
            for p in self.products:
                sku = (p['SKU'] or '').lower()
                nombre = (p['Nombre'] or '').lower()
                if query in sku or query in nombre:
                    matches.append((p['SKU'], p['Nombre']))
            
            if len(matches) <= 5 and len(matches) > 0:
                match_text = " | ".join([f"{s}: {n}" for s, n in matches[:5]])
                self.status_label.config(text=f"Encontrados: {match_text}")
            elif len(matches) > 5:
                self.status_label.config(text=f"Encontrados: {len(matches)} coincidencias")
            else:
                self.status_label.config(text=f"No encontrado")
        else:
            if self.status_label and self.usuario_actual:
                self.actualizar_status()
        
        self.filter_products()
    
    def selectFirstMatch(self, event=None):
        query = self.buscar_var.get().strip().lower()
        if not query:
            return
        
        for p in self.products:
            sku = (p['SKU'] or '').lower()
            nombre = (p['Nombre'] or '').lower()
            if query in sku or query in nombre:
                self.buscar_var.set(p['SKU'])
                self.filter_products()
                children = self.tree.get_children()
                for i, child in enumerate(children):
                    if self.tree.item(child)['values'][0] == p['SKU']:
                        self.tree.selection_remove(*self.tree.selection())
                        self.tree.selection_add(child)
                        self.tree.see(child)
                        return
        
        self.filter_products()
    
    def filter_products(self):
        self.tree.delete(*self.tree.get_children())
        
        buscar = self.buscar_var.get().lower()
        deposito = self.deposito_var.get()
        rubro = self.rubro_var.get()
        
        bajo_count = 0
        
        for p in self.products:
            if buscar and buscar not in (p['Nombre'] or '').lower() and buscar not in (p['SKU'] or '').lower():
                continue
            if deposito and p['deposito'] != deposito:
                continue
            if rubro and rubro != 'Todos' and p['Rubro'] != rubro:
                continue
            
            try:
                stock = int(p.get('stock', 0) or 0)
            except:
                stock = 0
            try:
                minimo = int(p.get('Stock Min', 0) or 0)
            except:
                minimo = 0
            
            tag = 'ok'
            if minimo and stock < minimo:
                bajo_count += 1
                tag = 'bajo'
            
            self.tree.insert('', tk.END, values=(
                p['SKU'] or '', p['Nombre'] or '', stock, minimo,
                p['deposito'] or '', p['precio'] or '', p['Estado'] or '',
            ), tags=(tag,))
        
        self.total_label.config(text=str(len(self.products)))
        self.bajo_label.config(text=str(bajo_count))
        
        rubros = sorted(set(p['Rubro'] for p in self.products if p['Rubro']))
        self.rubro_combo['values'] = ['Todos'] + rubros
    
    def add_product(self):
        self.open_editor()
    
    def edit_product(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Seleccionar", "Selecciona un producto")
            return
        
        sku = self.tree.item(sel[0])['values'][0]
        for p in self.products:
            if p['SKU'] == sku:
                self.open_editor(p)
                return
    
    def open_editor(self, product=None):
        win = tk.Toplevel(self.root)
        win.title("Editar" if product else "Agregar Producto")
        win.geometry("500x550")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        campos = [
            ('Nombre', 2), ('SKU', 3), ('Tipo (P/S)', 4), ('Estado (A/I)', 5),
            ('Moneda', 6), ('Rubro', 8), ('Subrubro', 9), ('Descripcion', 10),
            ('Precio', 16), ('Tasa IVA', 17), ('Costo', 18), 
            ('Stock Min', 21), ('Deposito', 24), ('Stock', 25)
        ]
        
        entries = {}
        
        for i, (label, col) in enumerate(campos):
            frame = ttk.Frame(win)
            frame.pack(fill=tk.X, padx=10, pady=3)
            ttk.Label(frame, text=label, width=15).pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=30)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            entries[col] = entry
        
        def save():
            row = self.ws_productos.max_row + 1
            
            for col, entry in entries.items():
                val = entry.get() or None
                if col in (21, 25) and val:
                    try:
                        val = int(val)
                    except:
                        val = 0
                self.ws_productos.cell(row, col).value = val
            
            self.ws_productos.cell(row, 1).value = self.ws_productos.cell(3, 1).value
            self.ws_productos.cell(row, 5).value = 'A'
            
            self.wb.save(EXCEL_FILE)
            win.destroy()
            self.init_excel()
            messagebox.showinfo("OK", "Producto guardado")
        
        ttk.Button(win, text="Guardar", command=save).pack(pady=15)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def delete_product(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Seleccionar", "Selecciona un producto")
            return
        
        if messagebox.askyesno("Eliminar", "Confirmar?"):
            sku = self.tree.item(sel[0])['values'][0]
            for p in self.products:
                if p['SKU'] == sku:
                    self.ws_productos.delete_rows(p['row'], 1)
                    self.wb.save(EXCEL_FILE)
                    self.init_excel()
                    messagebox.showinfo("OK", "Eliminado")
                    return
    
    def registrar_movimiento(self, sku, producto, tipo, cantidad, deposito, nota, nro_comp='', nro_factura=''):
        row = self.ws_movimientos.max_row + 1
        self.ws_movimientos.cell(row, 1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ws_movimientos.cell(row, 2).value = self.usuario_actual['nombre']
        self.ws_movimientos.cell(row, 3).value = sku
        self.ws_movimientos.cell(row, 4).value = producto
        self.ws_movimientos.cell(row, 5).value = tipo
        self.ws_movimientos.cell(row, 6).value = cantidad
        self.ws_movimientos.cell(row, 7).value = deposito
        self.ws_movimientos.cell(row, 8).value = nro_comp
        self.ws_movimientos.cell(row, 9).value = nro_factura
        self.ws_movimientos.cell(row, 10).value = nota
    
    def entrada_stock(self):
        win = tk.Toplevel(self.root)
        win.title("Entrada de Stock - Multiple")
        win.geometry("550x650")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        items_a_entrar = []
        
        datos_comp_frame = ttk.LabelFrame(win, text="Datos del Comprobante", padding=10)
        datos_comp_frame.pack(fill=tk.X, padx=10, pady=5)
        
        fecha_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Label(datos_comp_frame, text="Fecha:").pack(side=tk.LEFT)
        ttk.Entry(datos_comp_frame, textvariable=fecha_var, width=12).pack(side=tk.LEFT, padx=5)
        
        tipo_comp = tk.StringVar(value="Factura")
        ttk.Label(datos_comp_frame, text="Tipo:").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Combobox(datos_comp_frame, textvariable=tipo_comp, values=['Factura', 'Presupuesto', 'Remito'], width=10).pack(side=tk.LEFT, padx=5)
        
        nro_comp_var = tk.StringVar()
        ttk.Label(datos_comp_frame, text="Nro:").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(datos_comp_frame, textvariable=nro_comp_var, width=15).pack(side=tk.LEFT, padx=5)
        
        proveedor_frame = ttk.LabelFrame(win, text="Proveedor", padding=10)
        proveedor_frame.pack(fill=tk.X, padx=10, pady=5)
        
        proveedor_var = tk.StringVar()
        ttk.Label(proveedor_frame, text="CUIT:").pack(side=tk.LEFT)
        ttk.Entry(proveedor_frame, textvariable=proveedor_var, width=15).pack(side=tk.LEFT, padx=5)
        
        proveedor_nombre_var = tk.StringVar()
        proveedor_direccion_var = tk.StringVar()
        proveedor_telefono_var = tk.StringVar()
        
        def buscar_proveedor():
            cuits = proveedor_var.get().strip()
            if not cuits:
                return
            
            proveedores = self.get_proveedores()
            for prov in proveedores:
                if cuits in (prov['cuit'] or ''):
                    proveedor_nombre_var.set(prov['nombre'])
                    proveedor_direccion_var.set(prov.get('direccion', ''))
                    proveedor_telefono_var.set(prov.get('telefono', ''))
                    btn_guardar.config(command=lambda: guardar_manual(), text="Guardar en BD")
                    return
            
            result = self.buscar_arca(cuits)
            if result:
                proveedor_nombre_var.set(f"{result['nombre']}")
                proveedor_direccion_var.set(result.get('direccion', ''))
                proveedor_telefono_var.set(result.get('telefono', ''))
                
                def guardar_y_usar():
                    self.guardar_proveedor(result)
                    messagebox.showinfo("OK", "Proveedor guardado")
                
                btn_guardar.config(command=guardar_y_usar, text="Guardar de ARCA")
            else:
                proveedor_nombre_var.set("No encontrado - Cargar manualmente")
        
        def guardar_manual():
            cuits = proveedor_var.get().strip()
            if cuits and proveedor_nombre_var.get():
                self.guardar_proveedor({
                    'cuit': cuits,
                    'nombre': proveedor_nombre_var.get(),
                    'direccion': proveedor_direccion_var.get(),
                    'telefono': proveedor_telefono_var.get(),
                    'email': ''
                })
                messagebox.showinfo("OK", "Proveedor guardado manualmente")
        
        btn_buscar = ttk.Button(proveedor_frame, text="Buscar en ARCA", command=buscar_proveedor)
        btn_buscar.pack(side=tk.LEFT, padx=5)
        
        btn_guardar = ttk.Button(proveedor_frame, text="Guardar Manual", command=guardar_manual)
        btn_guardar.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(proveedor_frame, text="Nombre:").pack(side=tk.LEFT, padx=(15, 0))
        ttk.Entry(proveedor_frame, textvariable=proveedor_nombre_var, width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(proveedor_frame, text="Direccion:").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(proveedor_frame, textvariable=proveedor_direccion_var, width=20).pack(side=tk.LEFT, padx=5)
        
        producto_frame = ttk.LabelFrame(win, text="Buscar producto", padding=10)
        producto_frame.pack(fill=tk.X, padx=10, pady=5)
        
        sku_var = tk.StringVar()
        ttk.Label(producto_frame, text="SKU:").pack(side=tk.LEFT)
        sku_entry = ttk.Entry(producto_frame, textvariable=sku_var, width=20)
        sku_entry.pack(side=tk.LEFT, padx=5)
        sku_entry.focus()
        
        cantidad_var = tk.StringVar()
        ttk.Label(producto_frame, text="Cant:").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(producto_frame, textvariable=cantidad_var, width=8).pack(side=tk.LEFT, padx=5)
        
        costo_var = tk.StringVar()
        ttk.Label(producto_frame, text="Costo:").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(producto_frame, textvariable=costo_var, width=10).pack(side=tk.LEFT, padx=5)
        
        listbox = tk.Listbox(buscar_frame, height=5, width=35)
        listbox.pack(side=tk.LEFT, padx=5)
        
        def actualizar_lista(event=None):
            listbox.delete(0, tk.END)
            query = sku_var.get().strip().lower()
            if query:
                for p in self.products:
                    sku = (p['SKU'] or '').lower()
                    nombre = (p['Nombre'] or '').lower()
                    if query in sku or query in nombre:
                        listbox.insert(tk.END, f"{p['SKU']} - {p['Nombre']} (Stock: {p.get('stock', 0)})")
        
        sku_var.trace('w', lambda *a: actualizar_lista())
        
        def agregar_item():
            sel = listbox.curselection()
            if not sel:
                sku_check = sku_var.get().strip()
                for p in self.products:
                    if (p['SKU'] or '') == sku_check:
                        listbox.insert(tk.END, f"{p['SKU']} - {p['Nombre']} (Stock: {p.get('stock', 0)})")
                        sel = (0,)
                        break
                if not sel:
                    messagebox.showinfo("Nuevo SKU", "El SKU no existe. Use '+ Producto' para agregarlo.")
                    return
            
            try:
                cant = int(cantidad_var.get())
            except ValueError:
                messagebox.showerror("Error", "Cantidad inválida")
                return
            
            try:
                costo = float(costo_var.get()) if costo_var.get() else 0
            except:
                costo = 0
            
            txt = listbox.get(sel[0])
            sku = txt.split(" - ")[0]
            
            for p in self.products:
                if p['SKU'] == sku:
                    items_a_entrar.append({'sku': sku, 'nombre': p['Nombre'], 'cantidad': cant, 'deposito': p.get('deposito', 'Principal'), 'costo': costo})
                    break
            
            actualizar_lista_items()
            sku_var.set("")
            cantidad_var.set("")
            costo_var.set("")
        
        ttk.Button(producto_frame, text="+ Agregar", command=agregar_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(producto_frame, text="+ Nuevo Producto", command=lambda: self.add_product()).pack(side=tk.LEFT, padx=5)
        
        lista_frame = ttk.LabelFrame(win, text="Items a entrar", padding=10)
        lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ('SKU', 'Nombre', 'Cantidad')
        tree = ttk.Treeview(lista_frame, columns=cols, show='headings', height=8)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col != 'Nombre' else 200)
        tree.pack(fill=tk.BOTH, expand=True)
        
        def actualizar_lista_items():
            tree.delete(*tree.get_children())
            for item in items_a_entrar:
                tree.insert('', tk.END, values=(item['sku'], item['nombre'], item['cantidad']))
        
        def eliminar_item():
            sel = tree.selection()
            if sel:
                idx = tree.index(sel[0])
                items_a_entrar.pop(idx)
                actualizar_lista_items()
        
        ttk.Button(lista_frame, text="Eliminar selected", command=eliminar_item).pack(pady=5)
        
        datos_frame = ttk.LabelFrame(win, text="Datos del movimiento", padding=10)
        datos_frame.pack(fill=tk.X, padx=10, pady=5)
        
        nro_comp = tk.StringVar()
        ttk.Entry(datos_frame, textvariable=nro_comp, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(datos_frame, text="Nro Comp").pack(side=tk.LEFT)
        
        nro_fact = tk.StringVar()
        ttk.Entry(datos_frame, textvariable=nro_fact, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(datos_frame, text="Nro Fact").pack(side=tk.LEFT)
        
        nota = tk.StringVar()
        ttk.Entry(datos_frame, textvariable=nota, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(datos_frame, text="Nota").pack(side=tk.LEFT)
        
        def confirmar():
            if not items_a_entrar:
                messagebox.showwarning("警告", "No hay items")
                return
            
            for item in items_a_entrar:
                for p in self.products:
                    if p['SKU'] == item['sku']:
                        nuevo = p.get('stock', 0) + item['cantidad']
                        self.ws_productos.cell(p['row'], 25).value = nuevo
                        
                        row_lote = self.ws_movimientos.max_row + 1
                        self.ws_movimientos.cell(row_lote, 1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.ws_movimientos.cell(row_lote, 2).value = self.usuario_actual['nombre']
                        self.ws_movimientos.cell(row_lote, 3).value = item['sku']
                        self.ws_movimientos.cell(row_lote, 4).value = item['nombre']
                        self.ws_movimientos.cell(row_lote, 5).value = 'ENTRADA'
                        self.ws_movimientos.cell(row_lote, 6).value = item['cantidad']
                        self.ws_movimientos.cell(row_lote, 7).value = item['deposito']
                        self.ws_movimientos.cell(row_lote, 8).value = nro_comp.get()
                        self.ws_movimientos.cell(row_lote, 9).value = nro_fact.get()
                        self.ws_movimientos.cell(row_lote, 10).value = nota.get()
                        self.ws_movimientos.cell(row_lote, 11).value = item.get('costo', 0)
                        self.ws_movimientos.cell(row_lote, 12).value = item['cantidad']
                        break
            
            self.wb.save(EXCEL_FILE)
            win.destroy()
            self.init_excel()
            messagebox.showinfo("OK", f"{len(items_a_entrar)} entradas registradas (FIFO)")
        
        ttk.Button(win, text="Confirmar Todo", command=confirmar).pack(pady=15)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def salida_stock(self):
        win = tk.Toplevel(self.root)
        win.title("Salida de Stock - Multiple")
        win.geometry("500x550")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        items_a_salir = []
        
        buscar_frame = ttk.LabelFrame(win, text="Buscar producto", padding=10)
        buscar_frame.pack(fill=tk.X, padx=10, pady=5)
        
        buscar_var = tk.StringVar()
        ttk.Entry(buscar_frame, textvariable=buscar_var, width=25).pack(side=tk.LEFT, padx=5)
        
        cantidad_var = tk.StringVar()
        ttk.Entry(buscar_frame, textvariable=cantidad_var, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(buscar_frame, text="Cantidad").pack(side=tk.LEFT)
        
        listbox = tk.Listbox(buscar_frame, height=5, width=35)
        listbox.pack(side=tk.LEFT, padx=5)
        
        def actualizar_lista(event=None):
            listbox.delete(0, tk.END)
            query = buscar_var.get().strip().lower()
            if query:
                for p in self.products:
                    sku = (p['SKU'] or '').lower()
                    nombre = (p['Nombre'] or '').lower()
                    if query in sku or query in nombre:
                        listbox.insert(tk.END, f"{p['SKU']} - {p['Nombre']} (Stock: {p.get('stock', 0)})")
        
        buscar_var.trace('w', lambda *a: actualizar_lista())
        
        def agregar_item():
            sel = listbox.curselection()
            if not sel:
                return
            try:
                cant = int(cantidad_var.get())
            except ValueError:
                return
            
            txt = listbox.get(sel[0])
            sku = txt.split(" - ")[0]
            
            for p in self.products:
                if p['SKU'] == sku:
                    items_a_salir.append({'sku': sku, 'nombre': p['Nombre'], 'cantidad': cant, 'deposito': p.get('deposito', 'Principal'), 'stock': p.get('stock', 0)})
                    break
            
            actualizar_lista_items()
            buscar_var.set("")
            cantidad_var.set("")
        
        ttk.Button(buscar_frame, text="+ Agregar", command=agregar_item).pack(side=tk.LEFT, padx=5)
        
        lista_frame = ttk.LabelFrame(win, text="Items a salir", padding=10)
        lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ('SKU', 'Nombre', 'Cantidad', 'Stock Disp')
        tree = ttk.Treeview(lista_frame, columns=cols, show='headings', height=8)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=100 if col not in ('Nombre',) else 180)
        tree.pack(fill=tk.BOTH, expand=True)
        
        def actualizar_lista_items():
            tree.delete(*tree.get_children())
            for item in items_a_salir:
                tree.insert('', tk.END, values=(item['sku'], item['nombre'], item['cantidad'], item['stock']))
        
        def eliminar_item():
            sel = tree.selection()
            if sel:
                idx = tree.index(sel[0])
                items_a_salir.pop(idx)
                actualizar_lista_items()
        
        ttk.Button(lista_frame, text="Eliminar seleccionado", command=eliminar_item).pack(pady=5)
        
        datos_frame = ttk.LabelFrame(win, text="Datos del movimiento", padding=10)
        datos_frame.pack(fill=tk.X, padx=10, pady=5)
        
        nro_comp = tk.StringVar()
        ttk.Entry(datos_frame, textvariable=nro_comp, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(datos_frame, text="Nro Comp").pack(side=tk.LEFT)
        
        motivo = tk.StringVar()
        ttk.Combobox(datos_frame, textvariable=motivo, values=['Venta', 'Ajuste', 'Devolucion', 'Otro'], width=10).pack(side=tk.LEFT, padx=5)
        
        def confirmar():
            if not items_a_salir:
                return
            
            for item in items_a_salir:
                if item['cantidad'] > item['stock']:
                    messagebox.showerror("Error", f"Stock insuficiente para {item['sku']}")
                    return
            
            for item in items_a_salir:
                sku = item['sku']
                cantidad_necesaria = item['cantidad']
                
                lotes = self.get_lotes_fifo(sku)
                
                for lote in lotes:
                    if cantidad_necesaria <= 0:
                        break
                    
                    disponible = lote['disponible']
                    usar = min(cantidad_necesaria, disponible)
                    
                    row_lote = self.ws_movimientos.max_row + 1
                    self.ws_movimientos.cell(row_lote, 1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.ws_movimientos.cell(row_lote, 2).value = self.usuario_actual['nombre']
                    self.ws_movimientos.cell(row_lote, 3).value = sku
                    self.ws_movimientos.cell(row_lote, 4).value = item['nombre']
                    self.ws_movimientos.cell(row_lote, 5).value = 'SALIDA-FIFO'
                    self.ws_movimientos.cell(row_lote, 6).value = -usar
                    self.ws_movimientos.cell(row_lote, 7).value = item['deposito']
                    self.ws_movimientos.cell(row_lote, 8).value = nro_comp.get()
                    self.ws_movimientos.cell(row_lote, 9).value = ''
                    self.ws_movimientos.cell(row_lote, 10).value = motivo.get()
                    self.ws_movimientos.cell(row_lote, 11).value = lote['costo']
                    self.ws_movimientos.cell(row_lote, 12).value = f"Lote {lote['row']}"
                    
                    cantidad_necesaria -= usar
                
                for p in self.products:
                    if p['SKU'] == sku:
                        nuevo = p.get('stock', 0) - item['cantidad']
                        self.ws_productos.cell(p['row'], 25).value = nuevo
                        break
            
            self.wb.save(EXCEL_FILE)
            win.destroy()
            self.init_excel()
            messagebox.showinfo("OK", f"{len(items_a_salir)} salidas (FIFO)")
        
        ttk.Button(win, text="Confirmar Todo", command=confirmar).pack(pady=15)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def get_lotes_fifo(self, sku):
        lotes = []
        entrada_total = 0
        salida_total = 0
        
        for row in range(2, self.ws_movimientos.max_row + 1):
            row_sku = self.ws_movimientos.cell(row, 3).value
            if row_sku != sku:
                continue
            
            tipo = self.ws_movimientos.cell(row, 5).value
            cantidad = self.ws_movimientos.cell(row, 6).value or 0
            costo = self.ws_movimientos.cell(row, 11).value or 0
            
            if tipo == 'ENTRADA':
                entrada_total += cantidad
                lotes.append({'row': row, 'cantidad': cantidad, 'costo': costo, 'disponible': cantidad})
            elif tipo in ('SALIDA', 'SALIDA-FIFO'):
                salida_total += abs(cantidad)
        
        for lote in lotes:
            resta = min(salida_total, lote['disponible'])
            lote['disponible'] -= resta
            salida_total -= resta
        
        return [l for l in lotes if l['disponible'] > 0]
    
    def transferir(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Seleccionar", "Selecciona un producto")
            return
        
        item = self.tree.item(sel[0])['values']
        win = tk.Toplevel(self.root)
        win.title("Transferir")
        win.geometry("350x220")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        ttk.Label(win, text=f"Producto: {item[1]}").pack(pady=10)
        ttk.Label(win, text="Cantidad:").pack(pady=5)
        cantidad = tk.StringVar()
        ttk.Entry(win, textvariable=cantidad, width=15).pack()
        ttk.Label(win, text="Deposito destino:").pack(pady=5)
        destino = tk.StringVar()
        deps = [d for d in self.depositos if d != item[4]]
        ttk.Combobox(win, textvariable=destino, values=deps).pack()
        
        def confirmar():
            try:
                cant = int(cantidad.get())
            except ValueError:
                return
            
            if cant > item[2]:
                messagebox.showerror("Error", "Stock insuficiente")
                return
            
            sku = item[0]
            for p in self.products:
                if p['SKU'] == sku:
                    nuevo = p.get('stock', 0) - cant
                    self.ws_productos.cell(p['row'], 25).value = nuevo
                    self.registrar_movimiento(sku, item[1], 'TRANSFERENCIA', cant, destino.get(), f"De {item[4]}")
                    self.wb.save(EXCEL_FILE)
                    break
            
            win.destroy()
            self.init_excel()
            messagebox.showinfo("OK", "Transferencia realizada")
        
        ttk.Button(win, text="Confirmar", command=confirmar).pack(pady=15)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def historial(self):
        win = tk.Toplevel(self.root)
        win.title("Historial")
        win.geometry("900x450")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        ttk.Label(win, text="Historial de Movimientos", font=('', 14, 'bold')).pack(pady=15)
        
        cols = ('Fecha', 'Usuario', 'SKU', 'Producto', 'Tipo', 'Cantidad', 'Deposito', 'NroComp', 'NroFact', 'Nota')
        tree = ttk.Treeview(win, columns=cols, show='headings')
        
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=80 if col not in ('Producto', 'Nota') else 150)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        for row in range(2, self.ws_movimientos.max_row + 1):
            tree.insert('', tk.END, values=(
                self.ws_movimientos.cell(row, 1).value or '',
                self.ws_movimientos.cell(row, 2).value or '',
                self.ws_movimientos.cell(row, 3).value or '',
                self.ws_movimientos.cell(row, 4).value or '',
                self.ws_movimientos.cell(row, 5).value or '',
                self.ws_movimientos.cell(row, 6).value or '',
                self.ws_movimientos.cell(row, 7).value or '',
                self.ws_movimientos.cell(row, 8).value or '',
                self.ws_movimientos.cell(row, 9).value or '',
                self.ws_movimientos.cell(row, 10).value or '',
            ))
        
        ttk.Label(win, text=f"Total movimientos: {self.ws_movimientos.max_row - 1}").pack(pady=10)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def stock_bajo(self):
        win = tk.Toplevel(self.root)
        win.title("Stock Bajo")
        win.geometry("650x400")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        ttk.Label(win, text="Productos con Stock Bajo", font=('', 14, 'bold')).pack(pady=15)
        
        cols = ('SKU', 'Nombre', 'Stock', 'Stock Min', 'Faltante')
        tree = ttk.Treeview(win, columns=cols, show='headings')
        
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col != 'Nombre' else 180)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        count = 0
        for p in self.products:
            stock = int(p.get('stock', 0) or 0)
            minimo = int(p.get('Stock Min', 0) or 0)
            
            if minimo and stock < minimo:
                faltante = minimo - stock
                tree.insert('', tk.END, values=(p['SKU'], p['Nombre'], stock, minimo, faltante))
                count += 1
        
        ttk.Label(win, text=f"Total: {count} productos").pack(pady=10)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def reportes(self):
        win = tk.Toplevel(self.root)
        win.title("Reportes")
        win.geometry("700x450")
        
        theme = self.current_theme
        win.configure(bg=theme['bg'])
        
        ttk.Label(win, text="Reporte de Inventario", font=('', 14, 'bold')).pack(pady=15)
        
        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20)
        
        cols = ('SKU', 'Nombre', 'Stock', 'Costo', 'Total', 'Rubro')
        tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=90 if col != 'Nombre' else 170)
        
        tree.pack(fill=tk.BOTH, expand=True)
        
        total_valor = 0
        total_stock = 0
        
        for p in self.products:
            stock = int(p.get('stock', 0) or 0)
            costo = p.get('precio', 0) or 0
            total = stock * costo
            
            total_valor += total
            total_stock += stock
            
            tree.insert('', tk.END, values=(
                p['SKU'], p['Nombre'], stock,
                f"${costo:,.0f}" if costo else '-',
                f"${total:,.0f}",
                p['Rubro'] or ''
            ))
        
        ttk.Label(win, text=f"Unidades: {total_stock} | Valor: ${total_valor:,.0f}").pack(pady=10)
        tk.Label(win, text="Desarrollado por asubelzacg", font=('Arial', 8, 'bold'), fg=theme['watermark'], bg=theme['bg']).pack(pady=5)
    
    def guardar(self):
        self.wb.save(EXCEL_FILE)
        messagebox.showinfo("Guardado", "Excel guardado")

if __name__ == "__main__":
    root = tk.Tk()
    app = StockApp(root)
    root.mainloop()