import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

DEBUG_FILE = "debug.txt"

def log(msg):
    try:
        with open(DEBUG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()} - {msg}\n")
    except:
        pass

try:
    log("Starting...")
    
    EXCEL_FILE = "planilla_base.xlsx"
    
    LIGHT_THEME = {
        'bg': '#F5F7FA', 'fg': '#2C3E50', 'accent': '#3498DB',
        'toolbar': '#FFFFFF', 'tree_bg': '#FFFFFF', 'tree_fg': '#2C3E50',
        'tree_sel': '#3498DB', 'entry_bg': '#FFFFFF', 'entry_fg': '#2C3E50',
        'button': '#3498DB', 'watermark': '#7F8C8D', 'warning': '#E74C3C', 'success': '#27AE60'
    }
    
    USUARIOS_DEFAULT = [
        {'user': 'admin', 'pass': 'admin123', 'nombre': 'Administrador', 'rol': 'admin'},
        {'user': 'deposito', 'pass': 'depo123', 'nombre': 'Deposito', 'rol': 'user'},
        {'user': 'ventas', 'pass': 'vta123', 'nombre': 'Ventas', 'rol': 'user'},
    ]
    
    root = tk.Tk()
    root.title("Gestion de Stock")
    root.geometry("1100x750")
    
    wb = load_workbook(EXCEL_FILE)
    
    # Setup styles - THIS MIGHT FAIL
    log("Setting up styles")
    style = ttk.Style()
    try:
        style.theme_use('clam')
        log("Theme clam set")
    except Exception as e:
        log(f"Theme error: {e}")
    
    try:
        style.configure('TFrame', background=LIGHT_THEME['bg'])
        style.configure('TLabel', background=LIGHT_THEME['bg'], foreground=LIGHT_THEME['fg'])
        style.configure('Treeview', background=LIGHT_THEME['tree_bg'], foreground=LIGHT_THEME['tree_fg'], fieldbackground=LIGHT_THEME['tree_bg'])
        style.configure('Treeview.Heading', background=LIGHT_THEME['toolbar'], foreground=LIGHT_THEME['fg'])
        style.map('Treeview', background=[('selected', LIGHT_THEME['tree_sel'])])
        style.configure('TButton', background=LIGHT_THEME['button'], foreground='#FFFFFF')
        log("Styles configured")
    except Exception as e:
        log(f"Style config error: {e}")
    
    # Login
    login_win = tk.Toplevel(root)
    login_win.title("Login")
    login_win.geometry("300x220")
    login_win.configure(bg=LIGHT_THEME['bg'])
    
    ttk.Label(login_win, text="Gestion de Stock", font=('Arial', 14, 'bold')).pack(pady=20)
    user_var = tk.StringVar()
    ttk.Label(login_win, text="Usuario:").pack(pady=5)
    ttk.Entry(login_win, textvariable=user_var, width=20).pack()
    pass_var = tk.StringVar()
    ttk.Label(login_win, text="Contrasena:").pack(pady=5)
    pass_entry = ttk.Entry(login_win, textvariable=pass_var, show="*", width=20)
    pass_entry.pack()
    
    def entrar():
        user = user_var.get().strip()
        password = pass_var.get().strip()
        if 'Usuarios' not in wb.sheetnames:
            ws = wb.create_sheet('Usuarios')
            for col, h in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
                ws.cell(1, col).value = h
            for row, u in enumerate(USUARIOS_DEFAULT, 2):
                for col, key in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
                    ws.cell(row, col).value = u[key]
            wb.save(EXCEL_FILE)
        
        ws = wb['Usuarios']
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == user and ws.cell(row, 2).value == password:
                found = True
                break
        
        if found:
            login_win.destroy()
        else:
            messagebox.showerror("Error", "Usuario incorrecto")
    
    pass_entry.bind('<Return>', lambda e: entrar())
    ttk.Button(login_win, text="Ingresar", command=entrar).pack(pady=20)
    
    # Toolbar
    toolbar = tk.Frame(root, bg=LIGHT_THEME['toolbar'])
    toolbar.pack(side=tk.TOP, fill=tk.X)
    ttk.Button(toolbar, text="+ Producto", command=lambda: 0).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Entrada", command=lambda: 0).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Salida", command=lambda: 0).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Proveedores", command=lambda: 0).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Reportes", command=lambda: 0).pack(side=tk.LEFT, padx=2)
    
    # Filters
    filters = tk.Frame(root, bg=LIGHT_THEME['bg'])
    filters.pack(fill=tk.X, padx=10, pady=5)
    tk.Label(filters, text="Buscar:", bg=LIGHT_THEME['bg'], fg=LIGHT_THEME['fg']).pack(side=tk.LEFT)
    search_var = tk.StringVar()
    tk.Entry(filters, textvariable=search_var, width=30).pack(side=tk.LEFT, padx=5)
    
    # Filters 2 - comboboxes
    tipo_var = tk.StringVar(value="Todos")
    ttk.Label(filters, text="Tipo:").pack(side=tk.LEFT, padx=(20, 0))
    ttk.Combobox(filters, textvariable=tipo_var, values=['Todos', 'Producto', 'Insumo'], state='readonly', width=10).pack(side=tk.LEFT, padx=5)
    
    deposito_var = tk.StringVar(value="Todos")
    ttk.Label(filters, text="Deposito:").pack(side=tk.LEFT, padx=(20, 0))
    ttk.Combobox(filters, textvariable=deposito_var, values=['Todos', 'Principal'], state='readonly', width=10).pack(side=tk.LEFT, padx=5)
    
    # Treeview
    tree_frame = tk.Frame(root)
    tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    cols = ('SKU', 'Nombre', 'Stock', 'Costo', 'Deposito')
    tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.pack(fill=tk.BOTH, expand=True)
    
    # Status
    status_var = tk.StringVar(value="Listo")
    tk.Label(root, textvariable=status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X)
    
    log("Mainloop")
    root.mainloop()

except Exception as e:
    import traceback
    log(f"ERROR: {traceback.format_exc()}")