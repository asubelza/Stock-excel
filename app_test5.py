import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

DEBUG_FILE = "debug.txt"

def log(msg):
    try:
        with open(DEBUG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()} - {msg}\n")
    except:
        pass

try:
    log("Starting...")
    
    EXCEL_FILE = "planilla_base.xlsx"
    
    LIGHT_THEME = {
        'bg': '#F5F7FA', 'fg': '#2C3E50', 'accent': '#3498DB',
        'toolbar': '#FFFFFF', 'tree_bg': '#FFFFFF', 'tree_fg': '#2C3E50',
        'tree_sel': '#3498DB', 'entry_bg': '#FFFFFF', 'entry_fg': '#2C3E50',
        'button': '#3498DB', 'watermark': '#7F8C8D', 'warning': '#E74C3C', 'success': '#27AE60'
    }
    
    USUARIOS_DEFAULT = [
        {'user': 'admin', 'pass': 'admin123', 'nombre': 'Administrador', 'rol': 'admin'},
        {'user': 'deposito', 'pass': 'depo123', 'nombre': 'Deposito', 'rol': 'user'},
        {'user': 'ventas', 'pass': 'vta123', 'nombre': 'Ventas', 'rol': 'user'},
    ]
    
    root = tk.Tk()
    root.title("Gestion de Stock")
    root.geometry("1100x750")
    
    wb = load_workbook(EXCEL_FILE)
    
    # Login
    login_win = tk.Toplevel(root)
    login_win.title("Login")
    login_win.geometry("300x220")
    login_win.configure(bg=LIGHT_THEME['bg'])
    
    ttk.Label(login_win, text="Gestion de Stock", font=('Arial', 14, 'bold')).pack(pady=20)
    user_var = tk.StringVar()
    ttk.Label(login_win, text="Usuario:").pack(pady=5)
    ttk.Entry(login_win, textvariable=user_var, width=20).pack()
    pass_var = tk.StringVar()
    ttk.Label(login_win, text="Contrasena:").pack(pady=5)
    pass_entry = ttk.Entry(login_win, textvariable=pass_var, show="*", width=20)
    pass_entry.pack()
    
    def entrar():
        user = user_var.get().strip()
        password = pass_var.get().strip()
        if 'Usuarios' not in wb.sheetnames:
            ws = wb.create_sheet('Usuarios')
            for col, h in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
                ws.cell(1, col).value = h
            for row, u in enumerate(USUARIOS_DEFAULT, 2):
                for col, key in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
                    ws.cell(row, col).value = u[key]
            wb.save(EXCEL_FILE)
        
        ws = wb['Usuarios']
        usuario_actual = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == user and ws.cell(row, 2).value == password:
                usuario_actual = {'user': user, 'nombre': ws.cell(row, 3).value, 'rol': ws.cell(row, 4).value}
                break
        
        if usuario_actual:
            login_win.destroy()
        else:
            messagebox.showerror("Error", "Usuario incorrecto")
    
    pass_entry.bind('<Return>', lambda e: entrar())
    ttk.Button(login_win, text="Ingresar", command=entrar).pack(pady=20)
    
    # Toolbar with buttons
    toolbar = tk.Frame(root, bg=LIGHT_THEME['toolbar'])
    toolbar.pack(side=tk.TOP, fill=tk.X)
    
    log("Adding toolbar buttons")
    ttk.Button(toolbar, text="+ Producto", command=lambda: messagebox.showinfo("OK", "Add")).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Entrada", command=lambda: messagebox.showinfo("OK", "Entry")).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Salida", command=lambda: messagebox.showinfo("OK", "Exit")).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Proveedores", command=lambda: messagebox.showinfo("OK", "Prov")).pack(side=tk.LEFT, padx=2)
    ttk.Button(toolbar, text="Reportes", command=lambda: messagebox.showinfo("OK", "Rep")).pack(side=tk.LEFT, padx=2)
    
    # Theme toggle
    theme_var = tk.StringVar(value="Tema Claro")
    ttk.Button(toolbar, textvariable=theme_var, command=lambda: messagebox.showinfo("OK", "Theme")).pack(side=tk.RIGHT, padx=5)
    
    # Search
    search_frame = tk.Frame(root, bg=LIGHT_THEME['bg'])
    search_frame.pack(fill=tk.X, padx=10, pady=5)
    search_var = tk.StringVar()
    tk.Label(search_frame, text="Buscar:", bg=LIGHT_THEME['bg'], fg=LIGHT_THEME['fg']).pack(side=tk.LEFT)
    tk.Entry(search_frame, textvariable=search_var, width=30).pack(side=tk.LEFT, padx=5)
    
    # Treeview
    tree_frame = tk.Frame(root)
    tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    cols = ('SKU', 'Nombre', 'Stock', 'Costo', 'Deposito')
    tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.pack(fill=tk.BOTH, expand=True)
    
    # Status
    status_var = tk.StringVar(value="Listo")
    tk.Label(root, textvariable=status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X)
    
    log("Mainloop")
    root.mainloop()

except Exception as e:
    import traceback
    log(f"ERROR: {traceback.format_exc()}")