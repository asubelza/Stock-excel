import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
import os
import sys
from datetime import datetime

EXCEL_FILE = "planilla_base.xlsx"

if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
    excel_path = os.path.join(base_dir, EXCEL_FILE)
    if os.path.exists(excel_path):
        EXCEL_FILE = excel_path

LIGHT_THEME = {
    'bg': '#F5F7FA', 'fg': '#2C3E50', 'accent': '#3498DB',
    'toolbar': '#FFFFFF', 'tree_bg': '#FFFFFF', 'tree_fg': '#2C3E50',
    'tree_sel': '#3498DB', 'entry_bg': '#FFFFFF', 'entry_fg': '#2C3E50',
    'button': '#3498DB', 'watermark': '#7F8C8D', 'warning': '#E74C3C', 'success': '#27AE60'
}

DARK_THEME = {
    'bg': '#1A1A2E', 'fg': '#EAEAEA', 'accent': '#4A90D9',
    'toolbar': '#16213E', 'tree_bg': '#16213E', 'tree_fg': '#EAEAEA',
    'tree_sel': '#4A90D9', 'entry_bg': '#252545', 'entry_fg': '#EAEAEA',
    'button': '#4A90D9', 'watermark': '#6C7A89', 'warning': '#E74C3C', 'success': '#2ECC71'
}

USUARIOS_DEFAULT = [
    {'user': 'admin', 'pass': 'admin123', 'nombre': 'Administrador', 'rol': 'admin'},
    {'user': 'deposito', 'pass': 'depo123', 'nombre': 'Deposito', 'rol': 'user'},
    {'user': 'ventas', 'pass': 'vta123', 'nombre': 'Ventas', 'rol': 'user'},
]

dark_mode = False
current_theme = LIGHT_THEME

root = tk.Tk()
root.title("Gestion de Stock")
root.geometry("1100x750")

wb = load_workbook(EXCEL_FILE)

style = ttk.Style()
style.theme_use('clam')

def update_theme():
    global dark_mode, current_theme
    theme = DARK_THEME if dark_mode else LIGHT_THEME
    current_theme = theme
    root.configure(bg=theme['bg'])
    style.configure('TFrame', background=theme['bg'])
    style.configure('TLabel', background=theme['bg'], foreground=theme['fg'])
    style.configure('Treeview', background=theme['tree_bg'], foreground=theme['tree_fg'], fieldbackground=theme['tree_bg'])
    style.configure('Treeview.Heading', background=theme['toolbar'], foreground=theme['fg'])
    style.map('Treeview', background=[('selected', theme['tree_sel'])])
    style.configure('TButton', background=theme['button'], foreground='#FFFFFF')

ws_productos = wb.active
products = []
depositos = ['Principal']

for row in range(4, ws_productos.max_row + 1):
    nombre = ws_productos.cell(row, 2).value
    sku = ws_productos.cell(row, 3).value
    if sku or nombre:
        stock = ws_productos.cell(row, 25).value or 0
        deposito = ws_productos.cell(row, 24).value or 'Principal'
        if deposito not in depositos:
            depositos.append(deposito)
        products.append({
            'row': row, 'Nombre': nombre, 'SKU': sku,
            'Tipo': ws_productos.cell(row, 4).value,
            'Estado': ws_productos.cell(row, 5).value,
            'Rubro': ws_productos.cell(row, 8).value,
            'Stock Min': ws_productos.cell(row, 21).value or 0,
            'stock': stock, 'deposito': deposito,
            'precio': ws_productos.cell(row, 16).value,
        })

login_win = tk.Toplevel(root)
login_win.title("Login")
login_win.geometry("300x220")
login_win.configure(bg=LIGHT_THEME['bg'])

ttk.Label(login_win, text="Gestion de Stock", font=('Arial', 14, 'bold')).pack(pady=20)
user_var = tk.StringVar()
ttk.Label(login_win, text="Usuario:").pack(pady=5)
ttk.Entry(login_win, textvariable=user_var, width=20).pack()
pass_var = tk.StringVar()
ttk.Label(login_win, text="Contrasena:").pack(pady=5)
pass_entry = ttk.Entry(login_win, textvariable=pass_var, show="*", width=20)
pass_entry.pack()

def entrar():
    user = user_var.get().strip()
    password = pass_var.get().strip()
    if 'Usuarios' not in wb.sheetnames:
        ws = wb.create_sheet('Usuarios')
        for col, h in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
            ws.cell(1, col).value = h
        for row, u in enumerate(USUARIOS_DEFAULT, 2):
            for col, key in enumerate(['user', 'pass', 'nombre', 'rol'], 1):
                ws.cell(row, col).value = u[key]
        wb.save(EXCEL_FILE)
    
    ws = wb['Usuarios']
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == user and ws.cell(row, 2).value == password:
            found = True
            break
    
    if found:
        login_win.destroy()
        root.deiconify()
    else:
        messagebox.showerror("Error", "Usuario incorrecto")

pass_entry.bind('<Return>', lambda e: entrar())
ttk.Button(login_win, text="Ingresar", command=entrar).pack(pady=20)

root.withdraw()

def add_product():
    win = tk.Toplevel(root)
    win.title("Agregar Producto")
    win.geometry("500x600")
    win.configure(bg=current_theme['bg'])
    
    frame = ttk.LabelFrame(win, text="Datos del Producto", padding=15)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    vars = {}
    labels = ['Nombre', 'SKU', 'Tipo', 'Estado', 'Rubro', 'Cod Barra', 'Marca', 'Costo', 'Precio', 'Stock Min', 'Stock', 'Deposito']
    for i, label in enumerate(labels):
        ttk.Label(frame, text=label + ":").grid(row=i, column=0, sticky=tk.W, pady=3)
        if label in ['Tipo']:
            vars[label] = tk.StringVar(value='Producto')
            ttk.Combobox(frame, textvariable=vars[label], values=['Producto', 'Insumo'], width=25).grid(row=i, column=1, pady=3)
        elif label in ['Estado']:
            vars[label] = tk.StringVar(value='A')
            ttk.Combobox(frame, textvariable=vars[label], values=['A', 'B', 'C'], width=25).grid(row=i, column=1, pady=3)
        else:
            vars[label] = tk.StringVar()
            ttk.Entry(frame, textvariable=vars[label], width=27).grid(row=i, column=1, pady=3)
    
    def guardar():
        row = ws_productos.max_row + 1
        ws_productos.cell(row, 1).value = '1'
        ws_productos.cell(row, 2).value = vars['Nombre'].get()
        ws_productos.cell(row, 3).value = vars['SKU'].get()
        ws_productos.cell(row, 4).value = vars['Tipo'].get()
        ws_productos.cell(row, 5).value = vars['Estado'].get()
        ws_productos.cell(row, 8).value = vars['Rubro'].get()
        ws_productos.cell(row, 9).value = vars['Cod Barra'].get()
        ws_productos.cell(row, 10).value = vars['Marca'].get()
        ws_productos.cell(row, 16).value = vars['Costo'].get()
        ws_productos.cell(row, 17).value = vars['Precio'].get()
        ws_productos.cell(row, 21).value = vars['Stock Min'].get()
        ws_productos.cell(row, 24).value = vars['Deposito'].get() or 'Principal'
        ws_productos.cell(row, 25).value = vars['Stock'].get() or 0
        wb.save(EXCEL_FILE)
        messagebox.showinfo("OK", "Producto guardado")
        win.destroy()
    
    ttk.Button(win, text="Guardar", command=guardar).pack(pady=10)

def edit_product():
    sel = tree.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un producto")
        return
    
    item = tree.item(sel[0])
    sku = item['values'][0]
    
    for p in products:
        if p['SKU'] == sku:
            break
    
    win = tk.Toplevel(root)
    win.title("Editar Producto")
    win.geometry("500x600")
    win.configure(bg=current_theme['bg'])
    
    frame = ttk.LabelFrame(win, text="Datos del Producto", padding=15)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    vars = {}
    labels = ['Nombre', 'SKU', 'Tipo', 'Estado', 'Rubro', 'Cod Barra', 'Marca', 'Costo', 'Precio', 'Stock Min', 'Stock', 'Deposito']
    defaults = [p['Nombre'], p['SKU'], p['Tipo'], p['Estado'], p['Rubro'], '', '', str(p['precio'] or ''), '', str(p['Stock Min']), str(p['stock']), p['deposito']]
    
    for i, label in enumerate(labels):
        ttk.Label(frame, text=label + ":").grid(row=i, column=0, sticky=tk.W, pady=3)
        vars[label] = tk.StringVar(value=defaults[i])
        ttk.Entry(frame, textvariable=vars[label], width=27).grid(row=i, column=1, pady=3)
    
    def guardar():
        for col, label in enumerate(labels, 2):
            ws_productos.cell(p['row'], col).value = vars[label].get()
        wb.save(EXCEL_FILE)
        messagebox.showinfo("OK", "Producto actualizado")
        win.destroy()
    
    ttk.Button(win, text="Guardar", command=guardar).pack(pady=10)

def delete_product():
    sel = tree.selection()
    if not sel:
        messagebox.showwarning("Aviso", "Seleccione un producto")
        return
    
    if messagebox.askyesno("Confirmar", "Eliminar producto?"):
        item = tree.item(sel[0])
        sku = item['values'][0]
        for p in products:
            if p['SKU'] == sku:
                ws_productos.delete_rows(p['row'], 1)
                wb.save(EXCEL_FILE)
                break
        actualizar_lista()
        messagebox.showinfo("OK", "Producto eliminado")

def entrada_stock():
    win = tk.Toplevel(root)
    win.title("Entrada de Stock")
    win.geometry("600x700")
    win.configure(bg=current_theme['bg'])
    
    items = []
    
    datos_frame = ttk.LabelFrame(win, text="Datos del Comprobante", padding=10)
    datos_frame.pack(fill=tk.X, padx=10, pady=5)
    
    fecha_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
    ttk.Label(datos_frame, text="Fecha:").pack(side=tk.LEFT)
    ttk.Entry(datos_frame, textvariable=fecha_var, width=12).pack(side=tk.LEFT, padx=5)
    
    tipo_var = tk.StringVar(value="Factura")
    ttk.Label(datos_frame, text="Tipo:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Combobox(datos_frame, textvariable=tipo_var, values=['Factura', 'Presupuesto', 'Remito'], width=10).pack(side=tk.LEFT, padx=5)
    
    nro_comp_var = tk.StringVar()
    ttk.Label(datos_frame, text="Nro Comp:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(datos_frame, textvariable=nro_comp_var, width=15).pack(side=tk.LEFT, padx=5)
    
    prov_frame = ttk.LabelFrame(win, text="Proveedor", padding=10)
    prov_frame.pack(fill=tk.X, padx=10, pady=5)
    
    prov_cuit_var = tk.StringVar()
    prov_nombre_var = tk.StringVar()
    prov_direc_var = tk.StringVar()
    prov_tel_var = tk.StringVar()
    
    ttk.Label(prov_frame, text="CUIT:").pack(side=tk.LEFT)
    prov_cuit_entry = ttk.Entry(prov_frame, textvariable=prov_cuit_var, width=15)
    prov_cuit_entry.pack(side=tk.LEFT, padx=5)
    
    def buscar_proveedor():
        cuits = prov_cuit_var.get().strip().replace("-", "").replace(" ", "")
        if not cuits:
            return
        
        if 'Proveedores' in wb.sheetnames:
            ws = wb['Proveedores']
            for row in range(2, ws.max_row + 1):
                if cuits in (ws.cell(row, 1).value or '').replace("-", "").replace(" ", ""):
                    prov_nombre_var.set(ws.cell(row, 2).value)
                    return
        
        messagebox.showinfo("Buscando", f"Buscando CUIT en ARCA: {cuits}")
        
        try:
            import urllib.request
            import ssl
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            url = f"https://www.afip.gob.ar/rcel/consultaPadron/padron.aspx?cuitEmpresa={cuits}"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, context=ctx, timeout=10) as response:
                html = response.read().decode('utf-8', errors='ignore')
                
                import re
                nombre_match = re.search(r'<span[^>]*id="ctl00_ContentPlaceHolder1_lblDenominacion"[^>]*>([^<]+)<', html)
                direc_match = re.search(r'<span[^>]*id="ctl00_ContentPlaceHolder1_lblDomicilio"[^>]*>([^<]+)<', html)
                
                if nombre_match:
                    prov_nombre_var.set(nombre_match.group(1).strip())
                if direc_match:
                    prov_direc_var.set(direc_match.group(1).strip())
        except Exception as e:
            messagebox.showwarning("ARCA", f"No se encontró en ARCA.-error: {str(e)[:50]}")
    
    def guardar_proveedor():
        cuits = prov_cuit_var.get().strip()
        if not cuits or not prov_nombre_var.get():
            messagebox.showerror("Error", "CUIT y Nombre requeridos")
            return
        
        if 'Proveedores' not in wb.sheetnames:
            ws = wb.create_sheet('Proveedores')
            for col, h in enumerate(['cuit', 'nombre', 'direccion', 'telefono'], 1):
                ws.cell(1, col).value = h
        else:
            ws = wb['Proveedores']
        
        row = ws.max_row + 1
        ws.cell(row, 1).value = cuits
        ws.cell(row, 2).value = prov_nombre_var.get()
        ws.cell(row, 3).value = prov_direc_var.get()
        ws.cell(row, 4).value = prov_tel_var.get()
        wb.save(EXCEL_FILE)
        messagebox.showinfo("OK", "Proveedor guardado")
    
    ttk.Button(prov_frame, text="Buscar ARCA", command=buscar_proveedor).pack(side=tk.LEFT, padx=5)
    btn_guardar = ttk.Button(prov_frame, text="Guardar", command=guardar_proveedor)
    btn_guardar.pack(side=tk.LEFT, padx=5)
    ttk.Button(prov_frame, text="+ Nuevo", command=lambda: proveedores()).pack(side=tk.LEFT, padx=5)
    
    ttk.Label(prov_frame, text="Nombre:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prov_frame, textvariable=prov_nombre_var, width=25).pack(side=tk.LEFT, padx=5)
    ttk.Label(prov_frame, text="Direccion:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prov_frame, textvariable=prov_direc_var, width=25).pack(side=tk.LEFT, padx=5)
    
    prod_frame = ttk.LabelFrame(win, text="Buscar producto", padding=10)
    prod_frame.pack(fill=tk.X, padx=10, pady=5)
    
    sku_var = tk.StringVar()
    ttk.Label(prod_frame, text="SKU:").pack(side=tk.LEFT)
    sku_entry = ttk.Entry(prod_frame, textvariable=sku_var, width=15)
    sku_entry.pack(side=tk.LEFT, padx=5)
    sku_entry.focus()
    
    cant_var = tk.StringVar()
    ttk.Label(prod_frame, text="Cant:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prod_frame, textvariable=cant_var, width=8).pack(side=tk.LEFT, padx=5)
    
    costo_var = tk.StringVar()
    ttk.Label(prod_frame, text="Costo:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prod_frame, textvariable=costo_var, width=10).pack(side=tk.LEFT, padx=5)
    
    listbox = tk.Listbox(prod_frame, height=5, width=40)
    listbox.pack(pady=5)
    
    def buscar_prod(event=None):
        listbox.delete(0, tk.END)
        query = sku_var.get().strip().lower()
        if query:
            for p in products:
                if query in (p['SKU'] or '').lower() or query in (p['Nombre'] or '').lower():
                    listbox.insert(tk.END, f"{p['SKU']} - {p['Nombre']} (Stock: {p['stock']})")
    
    sku_var.trace('w', lambda *a: buscar_prod())
    
    def agregar():
        sel = listbox.curselection()
        if not sel:
            return
        txt = listbox.get(sel[0])
        sku = txt.split(" - ")[0]
        try:
            cant = int(cant_var.get())
            costo = float(costo_var.get()) if costo_var.get() else 0
        except:
            messagebox.showerror("Error", "Cantidad o costo inválido")
            return
        
        for p in products:
            if p['SKU'] == sku:
                items.append({'sku': sku, 'nombre': p['Nombre'], 'cantidad': cant, 'costo': costo})
                actualizar_lista()
                break
        
        sku_var.set("")
        cant_var.set("")
        costo_var.set("")
    
    def actualizar_lista():
        tree.delete(*tree.get_children())
        total_cant = 0
        total_importe = 0
        for item in items:
            tree.insert('', tk.END, values=(item['sku'], item['nombre'], item['cantidad']))
            total_cant += item['cantidad']
            total_importe += item['cantidad'] * item['costo']
        total_cant_var.set(str(total_cant))
        total_impo_var.set(f"${total_importe:,.0f}")
    
    ttk.Button(prod_frame, text="+ Agregar", command=agregar).pack(side=tk.LEFT, padx=5)
    ttk.Button(prod_frame, text="+ Nuevo Producto", command=add_product).pack(side=tk.LEFT, padx=5)
    
    lista_frame = ttk.LabelFrame(win, text="Items", padding=10)
    lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    cols = ('SKU', 'Nombre', 'Cantidad')
    tree = ttk.Treeview(lista_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=150)
    tree.pack(fill=tk.BOTH, expand=True)
    
    totales_frame = ttk.LabelFrame(win, text="Totales", padding=10)
    totales_frame.pack(fill=tk.X, padx=10, pady=5)
    
    total_cant_var = tk.StringVar(value="0")
    ttk.Label(totales_frame, text="Total Unidades:").pack(side=tk.LEFT, padx=10)
    ttk.Label(totales_frame, textvariable=total_cant_var, font=('', 12, 'bold')).pack(side=tk.LEFT)
    
    total_impo_var = tk.StringVar(value="$0")
    ttk.Label(totales_frame, text="Total Importe:").pack(side=tk.LEFT, padx=20)
    ttk.Label(totales_frame, textvariable=total_impo_var, font=('', 12, 'bold')).pack(side=tk.LEFT)
    
    def confirmar():
        if not items:
            messagebox.showwarning("Aviso", "No hay items")
            return
        
        for item in items:
            for p in products:
                if p['SKU'] == item['sku']:
                    nuevo = p['stock'] + item['cantidad']
                    ws_productos.cell(p['row'], 25).value = nuevo
                    
                    if 'Movimientos' not in wb.sheetnames:
                        ws_mov = wb.create_sheet('Movimientos')
                        for col, h in enumerate(['Fecha', 'Usuario', 'SKU', 'Producto', 'Tipo', 'Cantidad', 'Deposito'], 1):
                            ws_mov.cell(1, col).value = h
                    else:
                        ws_mov = wb['Movimientos']
                    
                    row_lote = ws_mov.max_row + 1
                    ws_mov.cell(row_lote, 1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws_mov.cell(row_lote, 2).value = user_var.get()
                    ws_mov.cell(row_lote, 3).value = item['sku']
                    ws_mov.cell(row_lote, 4).value = item['nombre']
                    ws_mov.cell(row_lote, 5).value = 'ENTRADA'
                    ws_mov.cell(row_lote, 6).value = item['cantidad']
                    ws_mov.cell(row_lote, 7).value = 'Principal'
                    ws_mov.cell(row_lote, 8).value = nro_comp_var.get()
                    ws_mov.cell(row_lote, 11).value = item['costo']
                    break
        
        wb.save(EXCEL_FILE)
        win.destroy()
        messagebox.showinfo("OK", f"{len(items)} entradas registradas")
        actualizar_lista_main()
    
    ttk.Button(win, text="Confirmar Todo", command=confirmar).pack(pady=15)

def salir():
    win = tk.Toplevel(root)
    win.title("Salida de Stock")
    win.geometry("600x700")
    win.configure(bg=current_theme['bg'])
    
    items = []
    
    datos_frame = ttk.LabelFrame(win, text="Datos del Comprobante", padding=10)
    datos_frame.pack(fill=tk.X, padx=10, pady=5)
    
    fecha_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
    ttk.Label(datos_frame, text="Fecha:").pack(side=tk.LEFT)
    ttk.Entry(datos_frame, textvariable=fecha_var, width=12).pack(side=tk.LEFT, padx=5)
    
    tipo_var = tk.StringVar(value="Venta")
    ttk.Label(datos_frame, text="Tipo:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Combobox(datos_frame, textvariable=tipo_var, values=['Venta', 'Uso interno', 'Merma', 'Devolucion'], width=10).pack(side=tk.LEFT, padx=5)
    
    nro_comp_var = tk.StringVar()
    ttk.Label(datos_frame, text="Nro Comp:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(datos_frame, textvariable=nro_comp_var, width=15).pack(side=tk.LEFT, padx=5)
    
    cliente_frame = ttk.LabelFrame(win, text="Cliente (opcional)", padding=10)
    cliente_frame.pack(fill=tk.X, padx=10, pady=5)
    
    cliente_var = tk.StringVar()
    ttk.Label(cliente_frame, text="Nombre:").pack(side=tk.LEFT)
    ttk.Entry(cliente_frame, textvariable=cliente_var, width=25).pack(side=tk.LEFT, padx=5)
    ttk.Label(cliente_frame, text="CUIT:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(cliente_frame, textvariable=cliente_var, width=15).pack(side=tk.LEFT, padx=5)
    
    prod_frame = ttk.LabelFrame(win, text="Buscar producto", padding=10)
    prod_frame.pack(fill=tk.X, padx=10, pady=5)
    
    sku_var = tk.StringVar()
    ttk.Label(prod_frame, text="SKU:").pack(side=tk.LEFT)
    sku_entry = ttk.Entry(prod_frame, textvariable=sku_var, width=15)
    sku_entry.pack(side=tk.LEFT, padx=5)
    sku_entry.focus()
    
    cant_var = tk.StringVar()
    ttk.Label(prod_frame, text="Cant:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prod_frame, textvariable=cant_var, width=8).pack(side=tk.LEFT, padx=5)
    
    listbox = tk.Listbox(prod_frame, height=5, width=40)
    listbox.pack(pady=5)
    
    def buscar_prod(event=None):
        listbox.delete(0, tk.END)
        query = sku_var.get().strip().lower()
        if query:
            for p in products:
                if p['stock'] > 0 and (query in (p['SKU'] or '').lower() or query in (p['Nombre'] or '').lower()):
                    listbox.insert(tk.END, f"{p['SKU']} - {p['Nombre']} (Stock: {p['stock']})")
    
    sku_var.trace('w', lambda *a: buscar_prod())
    
    def agregar():
        sel = listbox.curselection()
        if not sel:
            return
        txt = listbox.get(sel[0])
        sku = txt.split(" - ")[0]
        try:
            cant = int(cant_var.get())
        except:
            messagebox.showerror("Error", "Cantidad inválida")
            return
        
        for p in products:
            if p['SKU'] == sku:
                if p['stock'] < cant:
                    messagebox.showerror("Error", "Stock insuficiente")
                    return
                items.append({'sku': sku, 'nombre': p['Nombre'], 'cantidad': cant, 'stock': p['stock']})
                actualizar_lista()
                break
        
        sku_var.set("")
        cant_var.set("")
    
    def actualizar_lista():
        tree.delete(*tree.get_children())
        total_cant = 0
        for item in items:
            tree.insert('', tk.END, values=(item['sku'], item['nombre'], item['cantidad']))
            total_cant += item['cantidad']
        total_cant_var.set(str(total_cant))
    
    ttk.Button(prod_frame, text="+ Agregar", command=agregar).pack(side=tk.LEFT, padx=5)
    
    lista_frame = ttk.LabelFrame(win, text="Items a sair", padding=10)
    lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    cols = ('SKU', 'Nombre', 'Cantidad')
    tree = ttk.Treeview(lista_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=150)
    tree.pack(fill=tk.BOTH, expand=True)
    
    def eliminar_item():
        sel = tree.selection()
        if sel:
            idx = tree.index(sel[0])
            items.pop(idx)
            actualizar_lista()
    
    ttk.Button(lista_frame, text="Eliminar selected", command=eliminar_item).pack(pady=5)
    
    totales_frame = ttk.LabelFrame(win, text="Totales", padding=10)
    totales_frame.pack(fill=tk.X, padx=10, pady=5)
    
    total_cant_var = tk.StringVar(value="0")
    ttk.Label(totales_frame, text="Total Unidades:").pack(side=tk.LEFT, padx=10)
    ttk.Label(totales_frame, textvariable=total_cant_var, font=('', 12, 'bold')).pack(side=tk.LEFT)
    
    def confirmar():
        if not items:
            messagebox.showwarning("Aviso", "No hay items")
            return
        
        for item in items:
            for p in products:
                if p['SKU'] == item['sku']:
                    if p['stock'] < item['cantidad']:
                        messagebox.showerror("Error", f"Stock insuficiente para {p['SKU']}")
                        return
                    nuevo = p['stock'] - item['cantidad']
                    ws_productos.cell(p['row'], 25).value = nuevo
                    
                    if 'Movimientos' not in wb.sheetnames:
                        ws_mov = wb.create_sheet('Movimientos')
                        for col, h in enumerate(['Fecha', 'Usuario', 'SKU', 'Producto', 'Tipo', 'Cantidad', 'Deposito'], 1):
                            ws_mov.cell(1, col).value = h
                    else:
                        ws_mov = wb['Movimientos']
                    
                    row_lote = ws_mov.max_row + 1
                    ws_mov.cell(row_lote, 1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws_mov.cell(row_lote, 2).value = user_var.get()
                    ws_mov.cell(row_lote, 3).value = item['sku']
                    ws_mov.cell(row_lote, 4).value = item['nombre']
                    ws_mov.cell(row_lote, 5).value = 'SALIDA'
                    ws_mov.cell(row_lote, 6).value = item['cantidad']
                    ws_mov.cell(row_lote, 7).value = p['deposito']
                    ws_mov.cell(row_lote, 8).value = nro_comp_var.get()
                    break
        
        wb.save(EXCEL_FILE)
        win.destroy()
        messagebox.showinfo("OK", f"{len(items)} salidas registradas")
        actualizar_lista_main()
    
    ttk.Button(win, text="Confirmar Todo", command=confirmar).pack(pady=15)

def transferir():
    win = tk.Toplevel(root)
    win.title("Transferir Stock")
    win.geometry("500x350")
    win.configure(bg=current_theme['bg'])
    
    prod_frame = ttk.LabelFrame(win, text="Datos", padding=10)
    prod_frame.pack(fill=tk.X, padx=10, pady=5)
    
    sku_var = tk.StringVar()
    ttk.Label(prod_frame, text="SKU:").pack(side=tk.LEFT)
    ttk.Entry(prod_frame, textvariable=sku_var, width=20).pack(side=tk.LEFT, padx=5)
    
    cant_var = tk.StringVar()
    ttk.Label(prod_frame, text="Cant:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prod_frame, textvariable=cant_var, width=8).pack(side=tk.LEFT, padx=5)
    
    origen_var = tk.StringVar(value="Principal")
    ttk.Label(prod_frame, text="Origen:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Combobox(prod_frame, textvariable=origen_var, values=depositos, width=12).pack(side=tk.LEFT, padx=5)
    
    destino_var = tk.StringVar()
    ttk.Label(prod_frame, text="Destino:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(prod_frame, textvariable=destino_var, width=12).pack(side=tk.LEFT, padx=5)
    
    def confirmar():
        sku = sku_var.get().strip()
        try:
            cant = int(cant_var.get())
        except:
            messagebox.showerror("Error", "Cantidad inválida")
            return
        
        for p in products:
            if p['SKU'] == sku:
                if p['stock'] < cant:
                    messagebox.showerror("Error", "Stock insuficiente")
                    return
                ws_productos.cell(p['row'], 24).value = destino_var.get()
                wb.save(EXCEL_FILE)
                win.destroy()
                messagebox.showinfo("OK", "Transferencia realizada")
                actualizar_lista_main()
                return
        
        messagebox.showerror("Error", "Producto no encontrado")
    
    ttk.Button(win, text="Confirmar", command=confirmar).pack(pady=20)

def proveedores():
    win = tk.Toplevel(root)
    win.title("Proveedores")
    win.geometry("600x400")
    win.configure(bg=current_theme['bg'])
    
    frame = ttk.LabelFrame(win, text="Agregar Proveedor", padding=10)
    frame.pack(fill=tk.X, padx=10, pady=5)
    
    cuit_var = tk.StringVar()
    ttk.Label(frame, text="CUIT:").pack(side=tk.LEFT)
    ttk.Entry(frame, textvariable=cuit_var, width=15).pack(side=tk.LEFT, padx=5)
    
    nombre_var = tk.StringVar()
    ttk.Label(frame, text="Nombre:").pack(side=tk.LEFT, padx=(15, 0))
    ttk.Entry(frame, textvariable=nombre_var, width=25).pack(side=tk.LEFT, padx=5)
    
    def guardar():
        if not cuit_var.get() or not nombre_var.get():
            messagebox.showerror("Error", "CUIT y Nombre requeridos")
            return
        
        if 'Proveedores' not in wb.sheetnames:
            ws = wb.create_sheet('Proveedores')
            for col, h in enumerate(['cuit', 'nombre', 'direccion', 'telefono'], 1):
                ws.cell(1, col).value = h
        else:
            ws = wb['Proveedores']
        
        row = ws.max_row + 1
        ws.cell(row, 1).value = cuit_var.get()
        ws.cell(row, 2).value = nombre_var.get()
        wb.save(EXCEL_FILE)
        messagebox.showinfo("OK", "Proveedor guardado")
        nombre_var.set("")
        cuit_var.set("")
    
    ttk.Button(frame, text="Guardar", command=guardar).pack(side=tk.LEFT, padx=10)
    
    lista_frame = ttk.LabelFrame(win, text="Lista", padding=10)
    lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    cols = ('CUIT', 'Nombre')
    tree = ttk.Treeview(lista_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=150)
    tree.pack(fill=tk.BOTH, expand=True)
    
    if 'Proveedores' in wb.sheetnames:
        ws = wb['Proveedores']
        for row in range(2, ws.max_row + 1):
            tree.insert('', tk.END, values=(ws.cell(row, 1).value, ws.cell(row, 2).value))

def reportes():
    win = tk.Toplevel(root)
    win.title("Reportes")
    win.geometry("600x450")
    win.configure(bg=current_theme['bg'])
    
    ttk.Label(win, text="Reportes de Stock", font=('Arial', 14, 'bold')).pack(pady=10)
    
    ttk.Button(win, text="Stock por Producto", command=lambda: messagebox.showinfo("Info", "Reporte en desarrollo")).pack(pady=5)
    ttk.Button(win, text="Movimientos por Fecha", command=lambda: messagebox.showinfo("Info", "Reporte en desarrollo")).pack(pady=5)
    ttk.Button(win, text="Valor de Stock", command=lambda: messagebox.showinfo("Info", "Reporte en desarrollo")).pack(pady=5)
    ttk.Button(win, text="Rotacion de Stock", command=lambda: messagebox.showinfo("Info", "Reporte en desarrollo")).pack(pady=5)

def historial():
    if 'Movimientos' not in wb.sheetnames:
        messagebox.showinfo("Info", "No hay movimientos")
        return
    
    win = tk.Toplevel(root)
    win.title("Historial de Movimientos")
    win.geometry("800x400")
    win.configure(bg=current_theme['bg'])
    
    frame = ttk.LabelFrame(win, text="Filtro", padding=10)
    frame.pack(fill=tk.X, padx=10, pady=5)
    
    sku_var = tk.StringVar()
    ttk.Label(frame, text="SKU:").pack(side=tk.LEFT)
    ttk.Entry(frame, textvariable=sku_var, width=15).pack(side=tk.LEFT, padx=5)
    
    def buscar():
        tree.delete(*tree.get_children())
        ws = wb['Movimientos']
        sku = sku_var.get().strip().lower()
        for row in range(2, ws.max_row + 1):
            if sku and sku not in (ws.cell(row, 3).value or '').lower():
                continue
            tree.insert('', tk.END, values=(
                ws.cell(row, 1).value,
                ws.cell(row, 3).value,
                ws.cell(row, 4).value,
                ws.cell(row, 5).value,
                ws.cell(row, 6).value
            ))
    
    ttk.Button(frame, text="Buscar", command=buscar).pack(side=tk.LEFT, padx=10)
    
    lista_frame = ttk.Frame(win, padding=10)
    lista_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    cols = ('Fecha', 'SKU', 'Producto', 'Tipo', 'Cantidad')
    tree = ttk.Treeview(lista_frame, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120)
    tree.pack(fill=tk.BOTH, expand=True)

def stock_bajo():
    win = tk.Toplevel(root)
    win.title("Stock Bajo")
    win.geometry("600x400")
    win.configure(bg=current_theme['bg'])
    
    ttk.Label(win, text="Productos con Stock Bajo", font=('Arial', 14, 'bold')).pack(pady=10)
    
    cols = ('SKU', 'Nombre', 'Stock', 'Stock Min', 'Diferencia')
    tree = ttk.Treeview(win, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    for p in products:
        if p['stock'] < (p['Stock Min'] or 0):
            diff = p['stock'] - (p['Stock Min'] or 0)
            tree.insert('', tk.END, values=(p['SKU'], p['Nombre'], p['stock'], p['Stock Min'], diff))

def toggle_theme():
    global dark_mode
    dark_mode = not dark_mode
    update_theme()
    btn_theme.config(text="Modo Claro" if dark_mode else "Modo Oscuro")

toolbar = tk.Frame(root, bg=current_theme['toolbar'])
toolbar.pack(side=tk.TOP, fill=tk.X)

ttk.Button(toolbar, text="+ Producto", command=add_product).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Editar", command=edit_product).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Eliminar", command=delete_product).pack(side=tk.LEFT, padx=2)
ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
ttk.Button(toolbar, text="Entrada", command=entrada_stock).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Salida", command=salir).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Transferir", command=transferir).pack(side=tk.LEFT, padx=2)
ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
ttk.Button(toolbar, text="Proveedores", command=proveedores).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Reportes", command=reportes).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Stock Bajo", command=stock_bajo).pack(side=tk.LEFT, padx=2)
ttk.Button(toolbar, text="Historial", command=historial).pack(side=tk.LEFT, padx=2)
ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
btn_theme = ttk.Button(toolbar, text="Modo Oscuro", command=toggle_theme)
btn_theme.pack(side=tk.LEFT, padx=2)

filters = tk.Frame(root, bg=current_theme['bg'])
filters.pack(fill=tk.X, padx=10, pady=5)
tk.Label(filters, text="Buscar:", bg=current_theme['bg'], fg=current_theme['fg']).pack(side=tk.LEFT)
search_var = tk.StringVar()
tk.Entry(filters, textvariable=search_var, width=30).pack(side=tk.LEFT, padx=5)

tipo_var = tk.StringVar(value="Todos")
ttk.Label(filters, text="Tipo:").pack(side=tk.LEFT, padx=(20, 0))
ttk.Combobox(filters, textvariable=tipo_var, values=['Todos', 'Producto', 'Insumo'], state='readonly', width=10).pack(side=tk.LEFT, padx=5)

deposito_var = tk.StringVar(value="Todos")
ttk.Label(filters, text="Deposito:").pack(side=tk.LEFT, padx=(20, 0))
ttk.Combobox(filters, textvariable=deposito_var, values=['Todos'] + depositos, state='readonly', width=10).pack(side=tk.LEFT, padx=5)

tree_frame = tk.Frame(root)
tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
cols = ('SKU', 'Nombre', 'Stock', 'Costo', 'Deposito')
tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
for col in cols:
    tree.heading(col, text=col)
    tree.column(col, width=120)
tree.pack(fill=tk.BOTH, expand=True)

def actualizar_lista_main():
    global products, depositos
    products = []
    depositos = ['Principal']
    for row in range(4, ws_productos.max_row + 1):
        nombre = ws_productos.cell(row, 2).value
        sku = ws_productos.cell(row, 3).value
        if sku or nombre:
            stock = ws_productos.cell(row, 25).value or 0
            deposito = ws_productos.cell(row, 24).value or 'Principal'
            if deposito not in depositos:
                depositos.append(deposito)
            products.append({
                'row': row, 'Nombre': nombre, 'SKU': sku,
                'Tipo': ws_productos.cell(row, 4).value,
                'Estado': ws_productos.cell(row, 5).value,
                'Rubro': ws_productos.cell(row, 8).value,
                'Stock Min': ws_productos.cell(row, 21).value or 0,
                'stock': stock, 'deposito': deposito,
                'precio': ws_productos.cell(row, 16).value,
            })
    actualizar_lista()

def actualizar_lista():
    tree.delete(*tree.get_children())
    query = search_var.get().strip().lower()
    tipo = tipo_var.get()
    dep = deposito_var.get()
    
    for p in products:
        if query and query not in (p['SKU'] or '').lower() and query not in (p['Nombre'] or '').lower():
            continue
        if tipo != 'Todos' and p['Tipo'] != tipo:
            continue
        if dep != 'Todos' and p['deposito'] != dep:
            continue
        
        tree.insert('', tk.END, values=(p['SKU'], p['Nombre'], p['stock'], p['precio'], p['deposito']))

search_var.trace('w', lambda *a: actualizar_lista())
tipo_var.trace('w', lambda *a: actualizar_lista())
deposito_var.trace('w', lambda *a: actualizar_lista())

actualizar_lista()

status_var = tk.StringVar(value=f"Productos: {len(products)}")
tk.Label(root, textvariable=status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X)

root.mainloop()