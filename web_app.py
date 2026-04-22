from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flasgger import Swagger
from openpyxl import Workbook, load_workbook
import os
import logging
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.middleware.dispatcher import DispatcherMiddleware

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static')
app.secret_key = os.environ.get('SECRET_KEY', 'stock-secret-key-2024')

db_uri = os.environ.get('DATABASE_URL', 'sqlite:///stock.db')
if db_uri and db_uri.startswith('postgresql://'):
    db_uri = db_uri.replace('postgresql://', 'postgresql+psycopg2://')
app.config['SQLALCHEMY_DATABASE_URI'] = db_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_size': 10,
    'max_overflow': 20,
}

db = SQLAlchemy(app)
migrate = Migrate(app, db)

def usar_postgresql():
    return db_uri and 'postgresql' in db_uri.lower()

def lock_producto(sku):
    if usar_postgresql():
        import hashlib
        sku_hash = int(hashlib.md5(sku.encode()).hexdigest()[:16], 16) % (2**31)
        db.session.execute(db.text(f"SELECT pg_try_advisory_lock({sku_hash})"))

def unlock_producto(sku):
    if usar_postgresql():
        import hashlib
        sku_hash = int(hashlib.md5(sku.encode()).hexdigest()[:16], 16) % (2**31)
        db.session.execute(db.text(f"SELECT pg_advisory_unlock({sku_hash})"))

swagger = Swagger(app, template={
    "info": {
        "title": "Gestión de Stock API",
        "version": "1.0.0",
        "description": "API para gestión de stock, entradas, salidas, productos, proveedores, clientes y usuarios"
    },
    "definitions": {
        "UsuarioSchema": {
            "type": "object",
            "required": ["username", "password", "nombre"],
            "properties": {
                "username": {"type": "string"},
                "password": {"type": "string"},
                "nombre": {"type": "string"},
                "apellido": {"type": "string"},
                "rol": {"type": "string", "enum": ["admin", "datainput", "deposito"]}
            }
        },
        "ProductoSchema": {
            "type": "object",
            "required": ["sku", "nombre"],
            "properties": {
                "sku": {"type": "string"},
                "nombre": {"type": "string"},
                "stock": {"type": "integer"},
                "costo": {"type": "number"},
                "precio": {"type": "number"}
            }
        },
        "EntradaSchema": {
            "type": "object",
            "properties": {
                "nro_comp": {"type": "string"},
                "tipo_comp": {"type": "string"},
                "proveedor_cuit": {"type": "string"},
                "proveedor_nombre": {"type": "string"},
                "items": {"type": "array", "items": {"type": "object"}}
            }
        }
    }
})

USUARIOS = {
    'admin': {'pass': 'admin123', 'nombre': 'Administrador', 'rol': 'admin'},
    'deposito': {'pass': 'depo123', 'nombre': 'Deposito', 'rol': 'deposito'},
    'datainput': {'pass': 'data123', 'nombre': 'DataInput', 'rol': 'datainput'},
}

login_intentos = {}
MAX_INTENTOS = 5
BLOQUEO_MINUTOS = 15

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    apellido = db.Column(db.String(100))
    rol = db.Column(db.String(20), nullable=False, default='datainput')
    estado = db.Column(db.String(10), default='A')
    created_at = db.Column(db.DateTime, default=datetime.now)

PERMISOS = {
    'admin': ['/', '/historico', '/entrada', '/salida', '/proveedores', '/clientes', 
              '/nueva_entrada', '/nueva_salida', '/nuevo_producto', '/nuevo_proveedor',
              '/nuevo_cliente', '/importar_excel', '/exportar_excel', '/usuarios', '/logout',
              '/api/'],
    'datainput': ['/', '/historico', '/entrada', '/nueva_entrada', 
                  '/importar_excel', '/exportar_excel', '/logout', '/api/'],
    'deposito': ['/', '/historico', '/salida', '/nueva_salida', '/logout', '/api/'],
}

def tiene_permiso(ruta):
    rol = session.get('rol')
    permisos = PERMISOS.get(rol, [])
    return ruta in permisos

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'msg': 'No autenticado'}), 401
            return redirect('/stock/login')
        rol = session.get('rol')
        permisos = PERMISOS.get(rol, [])
        ruta = request.path
        
        if ruta in permisos:
            return f(*args, **kwargs)
        
        for p in permisos:
            if p != '/' and ruta.startswith(p):
                return f(*args, **kwargs)
        
        if request.path.startswith('/api/'):
            return jsonify({'ok': False, 'msg': 'Sin permisos'}), 403
        return redirect('/stock/')
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = request.form.get('usuario', '').strip()
        password = request.form.get('password', '').strip()
        
        if user in USUARIOS and USUARIOS[user]['pass'] == password:
            session['usuario'] = user
            session['nombre'] = USUARIOS[user]['nombre']
            session['rol'] = USUARIOS[user]['rol']
            return redirect('/stock/')
        
        db_user = Usuario.query.filter_by(username=user, estado='A').first()
        if db_user and db_user.password == password:
            session['usuario'] = db_user.username
            session['nombre'] = f"{db_user.nombre} {db_user.apellido or ''}".strip()
            session['rol'] = db_user.rol
            return redirect('/stock/')
        
        return render_template('login.html', error='Usuario o contraseña incorrectos')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/stock/login')

@app.route('/')
@login_required
def index():
    productos = Producto.query.order_by(Producto.nombre).all()
    return render_template('stock.html', productos=productos, usuario=session.get('nombre'))

@app.route('/entrada')
@login_required
def entrada():
    movimientos = Movimiento.query.filter_by(tipo='ENTRADA', eliminado=False).order_by(Movimiento.fecha.desc()).limit(100).all()
    return render_template('entrada.html', movimientos=movimientos, usuario=session.get('nombre'))

@app.route('/salida')
@login_required
def salida():
    movimientos = Movimiento.query.filter_by(tipo='SALIDA', eliminado=False).order_by(Movimiento.fecha.desc()).limit(100).all()
    return render_template('salida.html', movimientos=movimientos, usuario=session.get('nombre'))

@app.route('/historico')
@login_required
def historico():
    movimientos = Movimiento.query.filter_by(eliminado=False).order_by(Movimiento.fecha.desc()).limit(500).all()
    return render_template('historico.html', movimientos=movimientos, usuario=session.get('nombre'))

@app.route('/api/historico/limpiar', methods=['POST'])
@login_required
def api_limpiar_historico():
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'msg': 'Solo admins'}), 403
    try:
        eliminados = db.session.query(Movimiento).delete()
        db.session.commit()
        return jsonify({'ok': True, 'msg': f'{eliminados} movimientos eliminados'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/proveedores')
@login_required
def proveedores():
    proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
    return render_template('proveedores.html', proveedores=proveedores, usuario=session.get('nombre'))

@app.route('/clientes')
@login_required
def clientes():
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    return render_template('clientes.html', clientes=clientes, usuario=session.get('nombre'))

@app.route('/usuarios')
@login_required
def usuarios():
    if session.get('rol') != 'admin':
        return redirect('/stock/')
    usuarios = Usuario.query.order_by(Usuario.apellido, Usuario.nombre).all()
    return render_template('usuarios.html', usuarios=usuarios, usuario=session.get('nombre'))

@app.route('/nuevo_cliente')
@login_required
def nuevo_cliente():
    return render_template('nuevo_cliente.html', usuario=session.get('nombre'))

@app.route('/api/usuario', methods=['POST'])
@login_required
def api_usuario():
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'msg': 'Sin permisos'}), 403
    try:
        data = request.json
        if not data.get('username') or not data.get('password') or not data.get('nombre'):
            return jsonify({'ok': False, 'msg': 'Usuario, contraseña y nombre son requeridos'}), 400
        
        existing = Usuario.query.filter_by(username=data['username']).first()
        if existing:
            return jsonify({'ok': False, 'msg': 'Ya existe un usuario con ese nombre'}), 400
        
        usuario = Usuario(
            username=data['username'],
            password=data['password'],
            nombre=data['nombre'],
            apellido=data.get('apellido', ''),
            rol=data.get('rol', 'datainput')
        )
        db.session.add(usuario)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Usuario creado'})
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_usuario")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/usuario/<int:id>', methods=['DELETE'])
@login_required
def api_usuario_delete(id):
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'msg': 'Sin permisos'}), 403
    try:
        usuario = Usuario.query.get(id)
        if not usuario:
            return jsonify({'ok': False, 'msg': 'Usuario no encontrado'}), 404
        db.session.delete(usuario)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Usuario eliminado'})
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_usuario_delete")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/usuario/<int:id>', methods=['PUT'])
@login_required
def api_usuario_edit(id):
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'msg': 'Sin permisos'}), 403
    try:
        usuario = Usuario.query.get(id)
        if not usuario:
            return jsonify({'ok': False, 'msg': 'Usuario no encontrado'}), 404
        
        data = request.json
        
        if data.get('reset_password') and data.get('password'):
            usuario.password = data['password']
        
        if data.get('nombre'):
            usuario.nombre = data['nombre']
        if data.get('apellido'):
            usuario.apellido = data['apellido']
        if data.get('rol'):
            usuario.rol = data['rol']
        
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Usuario actualizado'})
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_usuario_edit")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/cliente', methods=['POST'])
@login_required
def api_cliente():
    try:
        data = request.json
        if not data.get('nombre'):
            return jsonify({'ok': False, 'msg': 'Nombre requerido'}), 400
        
        cliente = Cliente(
            nombre=data['nombre'],
            cuit=data.get('cuit', ''),
            direccion=data.get('direccion', ''),
            telefono=data.get('telefono', ''),
            email=data.get('email', '')
        )
        db.session.add(cliente)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Cliente creado'})
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_cliente")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/clientes')
@login_required
def api_clientes():
    query = request.args.get('q', '').strip()
    if query:
        q_lower = query.lower()
        clientes = Cliente.query.filter(
            db.or_(
                Cliente.nombre.ilike(f'%{q_lower}%'),
                Cliente.cuit.ilike(f'%{query}%')
            )
        ).limit(20).all()
    else:
        clientes = Cliente.query.limit(20).all()
    return jsonify([{
        'id': c.id, 'nombre': c.nombre, 'cuit': c.cuit or '',
        'direccion': c.direccion or '', 'telefono': c.telefono or ''
    } for c in clientes])

@app.route('/api/importar_clientes', methods=['POST'])
@login_required
def api_importar_clientes():
    try:
        file = request.files['archivo']
        if not file:
            return jsonify({'ok': False, 'msg': 'No hay archivo'}), 400
        
        wb = load_workbook(file)
        ws = wb.active
        
        creados = 0
        errores = []
        
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, 1).value
            cuit = ws.cell(row, 2).value
            
            if not nombre:
                continue
            
            try:
                cliente = Cliente(
                    nombre=str(nombre).strip(),
                    cuit=str(cuit).strip() if cuit else ''
                )
                db.session.add(cliente)
                creados += 1
            except Exception as e:
                errores.append(f'Fila {row}: {str(e)[:30]}')
        
        db.session.commit()
        msg = f'{creados} clientes importados'
        if errores:
            msg += f'. {len(errores)} errores'
        return jsonify({'ok': True, 'msg': msg})
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_importar_clientes")
        return jsonify({'ok': False, 'msg': str(e)}), 500

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(50), unique=True, nullable=False)
    nombre = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(20), default='P')
    estado = db.Column(db.String(10), default='A')
    rubro = db.Column(db.String(100))
    subrubro = db.Column(db.String(100))
    descripcion = db.Column(db.String(500))
    cod_proveedor = db.Column(db.String(50))
    observaciones = db.Column(db.String(500))
    precio = db.Column(db.Float, default=0)
    tasa_iva = db.Column(db.Float, default=21)
    costo = db.Column(db.Float, default=0)
    cod_barra = db.Column(db.String(50))
    stock_min = db.Column(db.Integer, default=0)
    stock = db.Column(db.Integer, default=0)
    deposito = db.Column(db.String(50), default='Principal')
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class Movimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.now)
    usuario = db.Column(db.String(50))
    sku = db.Column(db.String(50))
    producto = db.Column(db.String(200))
    tipo = db.Column(db.String(20), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    deposito = db.Column(db.String(50))
    nro_comp = db.Column(db.String(50))
    tipo_comp = db.Column(db.String(50))
    costo = db.Column(db.Float)
    proveedor_cuit = db.Column(db.String(20))
    proveedor_nombre = db.Column(db.String(200))
    cliente_cuit = db.Column(db.String(20))
    cliente_nombre = db.Column(db.String(200))
    observacion = db.Column(db.String(200))
    eliminado = db.Column(db.Boolean, default=False)
    eliminado_por = db.Column(db.String(50))
    eliminado_fecha = db.Column(db.DateTime)
    lote_id = db.Column(db.Integer)

class Proveedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    cuit = db.Column(db.String(20))
    direccion = db.Column(db.String(200))
    telefono = db.Column(db.String(50))
    email = db.Column(db.String(100))

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    cuit = db.Column(db.String(20))
    direccion = db.Column(db.String(200))
    telefono = db.Column(db.String(50))
    email = db.Column(db.String(100))

class Lote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(50), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    cantidad_disponible = db.Column(db.Integer, nullable=False)
    costo_unitario = db.Column(db.Float, default=0)
    fecha_ingreso = db.Column(db.DateTime, default=datetime.now)
    fecha_vencimiento = db.Column(db.DateTime)
    nro_lote = db.Column(db.String(50))
    deposito = db.Column(db.String(50))

COLUMNAS_EXCEL = {
    2: 'nombre', 3: 'sku', 4: 'tipo', 5: 'estado', 8: 'rubro',
    9: 'subrubro', 10: 'descripcion', 11: 'cod_proveedor', 15: 'observaciones',
    16: 'precio', 17: 'tasa_iva', 18: 'costo', 20: 'cod_barra',
    21: 'stock', 22: 'stock_min', 23: 'deposito'
}

with app.app_context():
    db.create_all()
    from sqlalchemy import inspect
    inspector = inspect(db.engine)
    
    try:
        columnas = [c['name'] for c in inspector.get_columns('movimiento')]
        if 'eliminado' not in columnas:
            db.session.execute(db.text("ALTER TABLE movimiento ADD COLUMN eliminado BOOLEAN DEFAULT FALSE"))
            db.session.commit()
        if 'eliminado_por' not in columnas:
            db.session.execute(db.text("ALTER TABLE movimiento ADD COLUMN eliminado_por VARCHAR(50)"))
            db.session.commit()
        if 'eliminado_fecha' not in columnas:
            db.session.execute(db.text("ALTER TABLE movimiento ADD COLUMN eliminado_fecha TIMESTAMP"))
            db.session.commit()
        if 'lote_id' not in columnas:
            db.session.execute(db.text("ALTER TABLE movimiento ADD COLUMN lote_id INTEGER"))
            db.session.commit()
    except Exception as e:
        print(f"Migration warning: {e}")

@app.route('/nueva_entrada')
@login_required
def nueva_entrada():
    return render_template('nueva_entrada.html', usuario=session.get('nombre'))

@app.route('/nueva_salida')
@login_required
def nueva_salida():
    return render_template('nueva_salida.html', usuario=session.get('nombre'))

@app.route('/nuevo_producto')
@login_required
def nuevo_producto():
    return render_template('nuevo_producto.html', usuario=session.get('nombre'))

@app.route('/nuevo_proveedor')
@login_required
def nuevo_proveedor():
    return render_template('nuevo_proveedor.html', usuario=session.get('nombre'))

@app.route('/api/productos')
@login_required
def api_productos():
    """
    Listar productos
    ---
    tags:
      - Productos
    parameters:
      - name: q
        in: query
        type: string
        required: false
        description: Buscar por SKU o nombre
    responses:
      200:
        description: Lista de productos
    """
    query = request.args.get('q', '').strip()
    if query:
        q_lower = query.lower()
        productos = Producto.query.filter(
            db.or_(
                Producto.sku.ilike(f'%{q_lower}%'),
                Producto.nombre.ilike(f'%{q_lower}%')
            )
        ).limit(20).all()
    else:
        productos = Producto.query.limit(20).all()
    return jsonify([{
        'sku': p.sku, 'nombre': p.nombre, 'stock': p.stock,
        'costo': p.costo or 0, 'deposito': p.deposito
    } for p in productos])

@app.route('/api/entrada', methods=['POST'])
@login_required
def api_entrada():
    """
    Registrar entrada de stock
    ---
    tags:
      - Entradas
    requestBody:
      required: true
      content:
        application/json:
          schema:
            type: object
            properties:
              nro_comp:
                type: string
              tipo_comp:
                type: string
              proveedor_cuit:
                type: string
              proveedor_nombre:
                type: string
              items:
                type: array
                items:
                  type: object
                  properties:
                    sku:
                      type: string
                    nombre:
                      type: string
                    cantidad:
                      type: integer
                    costo:
                      type: number
    responses:
      201:
        description: Entrada registrada
      400:
        description: Error
    """
    try:
        data = request.json
        items = data.get('items', [])
        
        if not items:
            return jsonify({'ok': False, 'msg': 'No hay items'}), 400
        
        for item in items:
            if not item.get('sku') or not item.get('cantidad'):
                continue
            
            producto = Producto.query.filter_by(sku=item['sku']).first()
            if not producto:
                producto = Producto(
                    sku=item['sku'],
                    nombre=item.get('nombre', item['sku']),
                    stock=0
                )
                db.session.add(producto)
            
            producto.stock = (producto.stock or 0) + item['cantidad']
            if item.get('costo'):
                producto.costo = item['costo']
            
            nuevo_lote = Lote(
                sku=item['sku'],
                cantidad=item['cantidad'],
                cantidad_disponible=item['cantidad'],
                costo_unitario=item.get('costo', 0),
                nro_lote=data.get('nro_comp', ''),
                deposito=item.get('deposito', 'Principal')
            )
            db.session.add(nuevo_lote)
            db.session.flush()
            
            movimiento = Movimiento(
                usuario=session.get('usuario', 'admin'),
                sku=item['sku'],
                producto=item.get('nombre', item['sku']),
                tipo='ENTRADA',
                cantidad=item['cantidad'],
                deposito=item.get('deposito', 'Principal'),
                nro_comp=data.get('nro_comp', ''),
                tipo_comp=data.get('tipo_comp', ''),
                costo=item.get('costo', 0),
                proveedor_cuit=data.get('proveedor_cuit', ''),
                lote_id=nuevo_lote.id,
                proveedor_nombre=data.get('proveedor_nombre', '')
            )
            db.session.add(movimiento)
        
        db.session.commit()
        return jsonify({'ok': True, 'msg': f'{len(items)} entradas registradas'})
    
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_entrada")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/salida', methods=['POST'])
@login_required
def api_salida():
    """
    Registrar salida de stock
    ---
    tags:
      - Salidas
    requestBody:
      required: true
      content:
        application/json:
          schema:
            type: object
            properties:
              nro_comp:
                type: string
              tipo_comp:
                type: string
              cliente_cuit:
                type: string
              cliente_nombre:
                type: string
              items:
                type: array
                items:
                  type: object
                  properties:
                    sku:
                      type: string
                    nombre:
                      type: string
                    cantidad:
                      type: integer
    responses:
      201:
        description: Salida registrada
      400:
        description: Error
    """
    try:
        data = request.json
        items = data.get('items', [])
        
        if not items:
            return jsonify({'ok': False, 'msg': 'No hay items'}), 400
        
        for item in items:
            if not item.get('sku') or not item.get('cantidad'):
                continue
            
            lock_producto(item['sku'])
            try:
                producto = Producto.query.filter_by(sku=item['sku']).first()
                if not producto:
                    return jsonify({'ok': False, 'msg': f'Producto {item["sku"]} no encontrado'}), 400
                
                stock_actual = producto.stock or 0
                if stock_actual < item['cantidad']:
                    return jsonify({'ok': False, 'msg': f'Stock insuficiente para {item["sku"]}. Stock actual: {stock_actual}, Solicitado: {item["cantidad"]}'}), 400
                
                cantidad_a_sacar = item['cantidad']
                lotes = Lote.query.filter_by(sku=item['sku']).filter(Lote.cantidad_disponible > 0).order_by(Lote.fecha_ingreso).all()
                
                lote_ids = []
                for lote in lotes:
                    if cantidad_a_sacar <= 0:
                        break
                    tomar = min(cantidad_a_sacar, lote.cantidad_disponible)
                    lote.cantidad_disponible -= tomar
                    cantidad_a_sacar -= tomar
                    lote_ids.append(f"{lote.nro_lote}:{tomar}")
                
                producto.stock -= item['cantidad']
                if producto.stock < 0:
                    db.session.rollback()
                    return jsonify({'ok': False, 'msg': f'Stock no puede ser negativo para {item["sku"]}. Stock: {producto.stock}'}), 400
                
                lote_ids_str = ", ".join(lote_ids)
                movimiento = Movimiento(
                    usuario=session.get('usuario', 'admin'),
                    sku=item['sku'],
                    producto=item.get('nombre', item['sku']),
                    tipo='SALIDA',
                    cantidad=item['cantidad'],
                    deposito=producto.deposito,
                    nro_comp=data.get('nro_comp', ''),
                    tipo_comp=data.get('tipo_comp', ''),
                    cliente_cuit=data.get('cliente_cuit', ''),
                    cliente_nombre=data.get('cliente_nombre', ''),
                    observacion=f"Lotes: {lote_ids_str}"
                )
                db.session.add(movimiento)
            finally:
                unlock_producto(item['sku'])
        
        db.session.commit()
        return jsonify({'ok': True, 'msg': f'{len(items)} salidas registradas'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/movimiento/<int:id>', methods=['PUT'])
@login_required
def api_movimiento_edit(id):
    try:
        movimiento = Movimiento.query.get(id)
        if not movimiento:
            return jsonify({'ok': False, 'msg': 'Movimiento no encontrado'}), 404
        
        rol = session.get('rol')
        if rol == 'datainput' and movimiento.tipo != 'ENTRADA':
            return jsonify({'ok': False, 'msg': 'Sin permisos para editar este tipo de movimiento'}), 403
        if rol == 'deposito' and movimiento.tipo != 'SALIDA':
            return jsonify({'ok': False, 'msg': 'Sin permisos para editar este tipo de movimiento'}), 403
        
        data = request.json
        cantidad_anterior = movimiento.cantidad
        nueva_cantidad = data.get('cantidad', movimiento.cantidad)
        
        if nueva_cantidad != cantidad_anterior:
            producto = Producto.query.filter_by(sku=movimiento.sku).first()
            if producto:
                diferencia = nueva_cantidad - cantidad_anterior
                stock_actual = producto.stock or 0
                
                if movimiento.tipo == 'ENTRADA':
                    producto.stock = stock_actual + diferencia
                    nuevo_lote = Lote(
                        sku=movimiento.sku,
                        cantidad=diferencia,
                        cantidad_disponible=diferencia,
                        costo_unitario=movimiento.costo or 0,
                        nro_lote=f"EDIT:{movimiento.id}",
                        deposito=producto.deposito
                    )
                    db.session.add(nuevo_lote)
                    db.session.flush()
                    movimiento.lote_id = nuevo_lote.id
                elif movimiento.tipo == 'SALIDA':
                    if diferencia > 0:
                        total_lote = sum(l.cantidad_disponible for l in Lote.query.filter_by(sku=movimiento.sku).all())
                        if total_lote < diferencia:
                            return jsonify({'ok': False, 'msg': f'Stock insuficiente en lotes. Disponible: {total_lote}, solicitado: {diferencia}'}), 400
                        cantidad_a_usar = diferencia
                        lotes = Lote.query.filter_by(sku=movimiento.sku).filter(Lote.cantidad_disponible > 0).order_by(Lote.fecha_ingreso).all()
                        for lote in lotes:
                            if cantidad_a_usar <= 0:
                                break
                            ocupar = min(cantidad_a_usar, lote.cantidad_disponible)
                            lote.cantidad_disponible -= ocupar
                            cantidad_a_usar -= ocupar
                    elif diferencia < 0:
                        restitucion = abs(diferencia)
                        lotes = Lote.query.filter_by(sku=movimiento.sku).order_by(Lote.fecha_ingreso.desc()).all()
                        for lote in lotes:
                            if restitucion <= 0:
                                break
                            lote.cantidad_disponible += restitucion
                            restitucion = 0
                if movimiento.tipo == 'SALIDA':
                    producto.stock = stock_actual - diferencia
                else:
                    producto.stock = stock_actual + diferencia
                    
                if producto.stock < 0:
                    db.session.rollback()
                    return jsonify({'ok': False, 'msg': f'Stock no puede ser negativo. Stock: {producto.stock}'}), 400
        
        movimiento.cantidad = nueva_cantidad
        
        if 'costo' in data:
            movimiento.costo = data['costo']
        if 'nro_comp' in data:
            movimiento.nro_comp = data['nro_comp']
        if 'tipo_comp' in data:
            movimiento.tipo_comp = data['tipo_comp']
        if 'proveedor_cuit' in data:
            movimiento.proveedor_cuit = data['proveedor_cuit']
        if 'proveedor_nombre' in data:
            movimiento.proveedor_nombre = data['proveedor_nombre']
        if 'cliente_cuit' in data:
            movimiento.cliente_cuit = data['cliente_cuit']
        if 'cliente_nombre' in data:
            movimiento.cliente_nombre = data['cliente_nombre']
        
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Movimiento actualizado'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/movimiento/<int:id>', methods=['DELETE'])
@login_required
def api_movimiento_delete(id):
    """Eliminar (anular) movimiento"""
    try:
        movimiento = Movimiento.query.get(id)
        if not movimiento:
            return jsonify({'ok': False, 'msg': 'Movimiento no encontrado'}), 404
        
        if movimiento.eliminado:
            return jsonify({'ok': False, 'msg': 'Movimiento ya fue eliminado'}), 400
        
        if session.get('rol') != 'admin':
            return jsonify({'ok': False, 'msg': 'Solo admins pueden eliminar movimientos'}), 403
        
        usuario = session.get('usuario', 'admin')
        
        producto = Producto.query.filter_by(sku=movimiento.sku).first()
        if producto:
            if movimiento.tipo == 'ENTRADA':
                nuevo_stock = (producto.stock or 0) - movimiento.cantidad
                if nuevo_stock < 0:
                    return jsonify({'ok': False, 'msg': f'No se puede eliminar. Stock quedaría en negativo: {nuevo_stock}'}), 400
                producto.stock = nuevo_stock
                if movimiento.lote_id:
                    lote = Lote.query.get(movimiento.lote_id)
                    if lote:
                        lote.cantidad_disponible = max(0, lote.cantidad_disponible - movimiento.cantidad)
            elif movimiento.tipo == 'SALIDA':
                producto.stock = (producto.stock or 0) + movimiento.cantidad
                cantidad_a_devolver = movimiento.cantidad
                lotes = Lote.query.filter_by(sku=movimiento.sku).order_by(Lote.fecha_ingreso.desc()).all()
                for lote in lotes:
                    if cantidad_a_devolver <= 0:
                        break
                    lote.cantidad_disponible += cantidad_a_devolver
                    cantidad_a_devolver = 0
        
        movimiento.eliminado = True
        movimiento.eliminado_por = usuario
        movimiento.eliminado_fecha = datetime.now()
        
        movimiento_anulado = Movimiento(
            usuario=usuario,
            sku=movimiento.sku,
            producto=movimiento.producto,
            tipo='ANULADO',
            cantidad=movimiento.cantidad,
            deposito=movimiento.deposito,
            nro_comp=movimiento.nro_comp,
            tipo_comp=movimiento.tipo_comp,
            costo=movimiento.costo,
            proveedor_cuit=movimiento.proveedor_cuit,
            proveedor_nombre=movimiento.proveedor_nombre,
            cliente_cuit=movimiento.cliente_cuit,
            cliente_nombre=movimiento.cliente_nombre,
            observacion=f"ANULADO - Original ID:{movimiento.id} - Usuario que anula:{usuario}"
        )
        db.session.add(movimiento_anulado)
        db.session.commit()
        
        return jsonify({'ok': True, 'msg': 'Movimiento anulado correctamente'})
    
    except Exception as e:
        db.session.rollback()
        logger.exception("Error en api_salida")
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/producto', methods=['POST'])
@login_required
def api_producto():
    """
    Crear nuevo producto
    ---
    tags:
      - Productos
    requestBody:
      required: true
      content:
        application/json:
          schema:
            type: object
            required:
              - sku
              - nombre
            properties:
              sku:
                type: string
              nombre:
                type: string
              stock:
                type: integer
              costo:
                type: number
              precio:
                type: number
    responses:
      201:
        description: Producto creado
      400:
        description: Error de validación
    """
    try:
        data = request.json
        if not data.get('sku') or not data.get('nombre'):
            return jsonify({'ok': False, 'msg': 'SKU y Nombre son requeridos'}), 400
        
        existing = Producto.query.filter_by(sku=data['sku']).first()
        if existing:
            return jsonify({'ok': False, 'msg': 'Ya existe un producto con ese SKU'}), 400
        
        producto = Producto(
            sku=data['sku'],
            nombre=data['nombre'],
            tipo=data.get('tipo', 'P'),
            estado=data.get('estado', 'A'),
            stock=data.get('stock', 0),
            stock_min=data.get('stock_min', 0),
            deposito=data.get('deposito', 'Principal'),
            precio=float(data.get('precio', 0) or 0),
            costo=float(data.get('costo', 0) or 0),
            rublo=data.get('rubro', ''),
            cod_barra=data.get('cod_barra', '')
        )
        db.session.add(producto)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Producto creado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/proveedores')
@login_required
def api_proveedores():
    """Listar proveedores"""
    query = request.args.get('q', '').strip()
    if query:
        q_lower = query.lower()
        proveedores = Proveedor.query.filter(
            db.or_(
                Proveedor.nombre.ilike(f'%{q_lower}%'),
                Proveedor.cuit.ilike(f'%{query}%')
            )
        ).limit(20).all()
    else:
        proveedores = Proveedor.query.limit(20).all()
    return jsonify([{
        'id': p.id, 'nombre': p.nombre, 'cuit': p.cuit or '',
        'direccion': p.direccion or '', 'telefono': p.telefono or ''
    } for p in proveedores])

@app.route('/api/proveedor', methods=['POST'])
@login_required
def api_proveedor():
    """
    Crear nuevo proveedor
    ---
    tags:
      - Proveedores
    requestBody:
      required: true
      content:
        application/json:
          schema:
            type: object
            required:
              - nombre
            properties:
              nombre:
                type: string
              cuit:
                type: string
              telefono:
                type: string
              email:
                type: string
              direccion:
                type: string
    responses:
      201:
        description: Proveedor creado
      400:
        description: Error
    """
    try:
        data = request.json
        if not data.get('nombre'):
            return jsonify({'ok': False, 'msg': 'Nombre requerido'}), 400
        
        proveedor = Proveedor(
            nombre=data['nombre'],
            cuit=data.get('cuit', ''),
            direccion=data.get('direccion', ''),
            telefono=data.get('telefono', ''),
            email=data.get('email', '')
        )
        db.session.add(proveedor)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'Proveedor creado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/api/importar_proveedores', methods=['POST'])
@login_required
def api_importar_proveedores():
    try:
        file = request.files['archivo']
        if not file:
            return jsonify({'ok': False, 'msg': 'No hay archivo'}), 400
        
        wb = load_workbook(file)
        ws = wb.active
        
        creados = 0
        errores = []
        
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, 1).value
            cuit = ws.cell(row, 2).value
            
            if not nombre:
                continue
            
            try:
                proveedor = Proveedor(
                    nombre=str(nombre).strip(),
                    cuit=str(cuit).strip() if cuit else ''
                )
                db.session.add(proveedor)
                creados += 1
            except Exception as e:
                errores.append(f'Fila {row}: {str(e)[:30]}')
        
        db.session.commit()
        msg = f'{creados} proveedores importados'
        if errores:
            msg += f'. {len(errores)} errores'
        return jsonify({'ok': True, 'msg': msg})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route('/importar_excel', methods=['GET', 'POST'])
@login_required
def importar_excel():
    if request.method == 'POST':
        try:
            file = request.files['archivo']
            if not file:
                return jsonify({'ok': False, 'msg': 'No hay archivo'}), 400
            
            wb = load_workbook(file)
            ws = wb.active
            
            productos_creados = 0
            productos_actualizados = 0
            errores = []
            
            for row_num in range(4, ws.max_row + 1):
                sku = ws.cell(row_num, 3).value
                nombre = ws.cell(row_num, 2).value
                
                if not sku or not nombre:
                    continue
                
                try:
                    existing = Producto.query.filter_by(sku=str(sku).strip()).first()
                    
                    producto_data = {}
                    for col, field in COLUMNAS_EXCEL.items():
                        val = ws.cell(row_num, col).value
                        if field == 'stock' and val is None:
                            val = 0
                        producto_data[field] = val
                    
                    if existing:
                        for field, val in producto_data.items():
                            if val is not None and hasattr(existing, field):
                                setattr(existing, field, val)
                        existing.updated_at = datetime.now()
                        productos_actualizados += 1
                    else:
                        nuevo = Producto(
                            sku=str(sku).strip(),
                            nombre=str(nombre).strip(),
                            tipo=producto_data.get('tipo', 'P'),
                            estado=producto_data.get('estado', 'A'),
                            rubro=producto_data.get('rubro', ''),
                            subrubro=producto_data.get('subrubro', ''),
                            descripcion=producto_data.get('descripcion', ''),
                            cod_proveedor=producto_data.get('cod_proveedor', ''),
                            observaciones=producto_data.get('observaciones', ''),
                            precio=float(producto_data.get('precio') or 0),
                            tasa_iva=float(producto_data.get('tasa_iva') or 21),
                            costo=float(producto_data.get('costo') or 0),
                            cod_barra=str(producto_data.get('cod_barra') or ''),
                            stock_min=int(producto_data.get('stock_min') or 0),
                            stock=int(producto_data.get('stock') or 0),
                            deposito=str(producto_data.get('deposito') or 'Principal')
                        )
                        db.session.add(nuevo)
                        productos_creados += 1
                
                except Exception as e:
                    import sys
                    print(f'ERROR fila {row_num}: SKU={sku} nombre={nombre} - {e}', file=sys.stderr)
                    errores.append(f'Fila {row_num} SKU={sku}: {str(e)[:80]}')
            
            db.session.commit()
            
            msg = f'Importación: {productos_creados} creados, {productos_actualizados} actualizados'
            if errores:
                msg += f'. Errores: {len(errores)}'
            
            return jsonify({'ok': True, 'msg': msg})
        
        except Exception as e:
            db.session.rollback()
            return jsonify({'ok': False, 'msg': f'Error: {str(e)}'}), 500
    
    return render_template('importar.html', usuario=session.get('nombre'))

@app.route('/exportar_excel')
@login_required
def exportar_excel():
    try:
        from flask import send_file
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        headers = ['Cod', 'Nombre', 'SKU', 'Tipo', 'Estado', 'Rubro', 'Subrubro', 'Descripcion',
                   'Cod Proveedor', 'Observaciones', 'Precio', 'Tasa IVA', 'Costo', 'Cod Barra',
                   'Stock Min', 'Deposito', 'Stock']
        for col, h in enumerate(headers, 1):
            ws.cell(1, col).value = h
        
        productos = Producto.query.order_by(Producto.nombre).all()
        for row, p in enumerate(productos, 2):
            ws.cell(row, 1).value = '1'
            ws.cell(row, 2).value = p.nombre
            ws.cell(row, 3).value = p.sku
            ws.cell(row, 4).value = p.tipo
            ws.cell(row, 5).value = p.estado
            ws.cell(row, 6).value = p.rubro
            ws.cell(row, 7).value = p.subrubro
            ws.cell(row, 8).value = p.descripcion
            ws.cell(row, 9).value = p.cod_proveedor
            ws.cell(row, 10).value = p.observaciones
            ws.cell(row, 11).value = p.precio
            ws.cell(row, 12).value = p.tasa_iva
            ws.cell(row, 13).value = p.costo
            ws.cell(row, 14).value = p.cod_barra
            ws.cell(row, 15).value = p.stock_min
            ws.cell(row, 16).value = p.deposito
            ws.cell(row, 17).value = p.stock
        
        filename = f'stock_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return f'Error: {str(e)}'

@app.route('/swagger.json')
def swagger_json():
    """Endpoint que retorna la especificación OpenAPI en JSON"""
    return jsonify(swagger.config.get('spec', {}))

if __name__ == '__main__':
    app.run(debug=os.environ.get('DEBUG', 'True').lower() == 'true', host='0.0.0.0', port=5000)

application = DispatcherMiddleware(app, {'/stock': app})